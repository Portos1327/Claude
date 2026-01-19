# -*- coding: utf-8 -*-

import base64
import io
from datetime import datetime
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import logging

_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportQuickdevisWizard(models.TransientModel):
    _name = 'export.quickdevis.wizard'
    _description = 'Assistant d\'export vers QuickDevis 7'

    article_ids = fields.Many2many('vst.article', string='Articles à exporter')
    
    export_all = fields.Boolean(string='Exporter tous les articles', default=False)
    include_deleted = fields.Boolean(string='Inclure les articles supprimés', default=False)
    
    famille_ids = fields.Many2many(
        'vst.famille',
        'export_qdv_vst_wizard_famille_rel',
        'wizard_id', 'famille_id',
        string='Familles',
        domain=[('parent_id', '=', False)]
    )
    
    export_file = fields.Binary(string='Fichier exporté', readonly=True)
    export_filename = fields.Char(string='Nom du fichier', readonly=True)
    state = fields.Selection([
        ('draft', 'Configuration'),
        ('done', 'Terminé')
    ], default='draft')
    articles_exported = fields.Integer(string='Articles exportés', readonly=True)

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if self.env.context.get('active_ids'):
            res['article_ids'] = [(6, 0, self.env.context.get('active_ids'))]
        return res

    def action_export(self):
        """Lance l'export vers QuickDevis 7"""
        self.ensure_one()
        
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("La bibliothèque openpyxl n'est pas installée."))
        
        # Déterminer les articles à exporter
        if self.export_all:
            domain = [('active', '=', True)]
            if not self.include_deleted:
                domain.append(('is_deleted', '=', False))
            articles = self.env['vst.article'].search(domain)
        elif self.famille_ids:
            all_famille_ids = self._get_all_subfamilies(self.famille_ids.ids)
            domain = [('famille_id', 'in', all_famille_ids), ('active', '=', True)]
            if not self.include_deleted:
                domain.append(('is_deleted', '=', False))
            articles = self.env['vst.article'].search(domain)
        elif self.article_ids:
            articles = self.article_ids
        else:
            raise UserError(_('Veuillez sélectionner des articles ou cocher "Exporter tous les articles".'))
        
        if not articles:
            raise UserError(_('Aucun article à exporter.'))
        
        try:
            wb = Workbook()
            
            # Onglet Articles
            ws_articles = wb.active
            ws_articles.title = "Articles VST"
            
            # En-têtes
            headers = [
                'Code Article', 'Désignation', 'Référence Fabricant', 'Nom Fabricant',
                'Prix Achat', 'Prix Public HT', 'Remise %', 'Unité',
                'Famille', 'Sous-famille', 'Marque', 'Activité',
                'Écotaxe HT', 'Type Article', 'Date Prix',
                'Code Famille 1', 'Code Famille 2', 'Code Famille 3',
                'Code Famille QDV7'
            ]
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col, header in enumerate(headers, 1):
                cell = ws_articles.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Données
            for row, article in enumerate(articles, 2):
                # Extraire les codes famille
                famille_parts = (article.nouvelle_famille_code or article.famille_code or '').split('.')
                code_f1 = famille_parts[0] if len(famille_parts) > 0 else ''
                code_f2 = famille_parts[1] if len(famille_parts) > 1 else ''
                code_f3 = famille_parts[2] if len(famille_parts) > 2 else ''
                
                ws_articles.cell(row, 1, article.code_article or '')
                ws_articles.cell(row, 2, article.designation or '')
                ws_articles.cell(row, 3, article.reference_fabricant or '')
                ws_articles.cell(row, 4, article.nom_fabricant or '')
                ws_articles.cell(row, 5, article.prix_achat_adherent or 0)
                ws_articles.cell(row, 6, article.prix_public_ht or 0)
                ws_articles.cell(row, 7, article.remise or 0)
                ws_articles.cell(row, 8, article.unite or 'U')
                ws_articles.cell(row, 9, article.libelle_famille or '')
                ws_articles.cell(row, 10, article.libelle_sous_famille or '')
                ws_articles.cell(row, 11, article.libelle_marque or '')
                ws_articles.cell(row, 12, article.libelle_activite or '')
                ws_articles.cell(row, 13, article.ecotaxe_ht or 0)
                ws_articles.cell(row, 14, article.type_article or '')
                ws_articles.cell(row, 15, str(article.date_dernier_prix) if article.date_dernier_prix else '')
                ws_articles.cell(row, 16, code_f1)
                ws_articles.cell(row, 17, code_f2)
                ws_articles.cell(row, 18, code_f3)
                # Formule pour Code Famille QDV7
                ws_articles.cell(row, 19, f'=P{row}&Q{row}&R{row}')
            
            # Ajuster les largeurs de colonnes
            for col in range(1, len(headers) + 1):
                ws_articles.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = 15
            
            # Onglet Structure (hiérarchie pour QDV7)
            ws_structure = wb.create_sheet("Structure")
            
            structure_headers = ['Code Famille', 'Libellé Famille', 'Code Sous-Famille', 'Libellé Sous-Famille']
            for col, header in enumerate(structure_headers, 1):
                cell = ws_structure.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Récupérer la hiérarchie des familles utilisées
            familles_used = set()
            for article in articles:
                if article.famille_id:
                    familles_used.add(article.famille_id.id)
            
            if familles_used:
                familles = self.env['vst.famille'].browse(list(familles_used))
                
                structure_data = {}
                for famille in familles:
                    if famille.parent_id:
                        parent_code = famille.parent_id.code or ''
                        parent_name = famille.parent_id.name or ''
                        child_code = famille.code or ''
                        child_name = famille.name or ''
                        
                        key = (parent_code, parent_name, child_code, child_name)
                        structure_data[key] = True
                
                row = 2
                for (parent_code, parent_name, child_code, child_name) in sorted(structure_data.keys()):
                    ws_structure.cell(row, 1, parent_code)
                    ws_structure.cell(row, 2, parent_name)
                    ws_structure.cell(row, 3, child_code)
                    ws_structure.cell(row, 4, child_name)
                    row += 1
            
            # Sauvegarder
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"Export_QuickDevis7_VST_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            self.write({
                'export_file': base64.b64encode(output.read()),
                'export_filename': filename,
                'state': 'done',
                'articles_exported': len(articles),
            })
            
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'export.quickdevis.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }
            
        except Exception as e:
            _logger.error("Erreur export QuickDevis: %s", str(e))
            raise UserError(_('Erreur lors de l\'export :\n%s') % str(e))

    def _get_all_subfamilies(self, family_ids):
        """Récupère récursivement toutes les sous-familles"""
        Family = self.env['vst.famille']
        all_ids = list(family_ids)
        children = Family.search([('parent_id', 'in', family_ids)])
        if children:
            all_ids.extend(self._get_all_subfamilies(children.ids))
        return all_ids

    def action_back(self):
        """Retour à l'étape de configuration"""
        self.write({
            'state': 'draft',
            'export_file': False,
            'export_filename': False,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.quickdevis.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
