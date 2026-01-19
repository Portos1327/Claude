# -*- coding: utf-8 -*-

import base64
import io
import os
from datetime import datetime
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import logging

_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportNaviwestWizard(models.TransientModel):
    _name = 'export.naviwest.wizard'
    _description = 'Assistant d\'export vers Naviwest'

    export_type = fields.Selection([
        ('beg', 'BEG'),
        ('niedax', 'NIEDAX'),
        ('cables', 'CÂBLES'),
    ], string='Type d\'export', required=True, default='beg')
    
    filter_field = fields.Selection([
        ('libelle_famille', 'Famille'),
        ('libelle_sous_famille', 'Sous-famille'),
        ('nom_fabricant', 'Fabricant'),
        ('libelle_marque', 'Marque'),
    ], string='Filtrer par')
    
    filter_value = fields.Char(string='Valeur du filtre', help='Ex: CABLE, LEGRAND, etc.')
    
    article_ids = fields.Many2many('vst.article', string='Articles à exporter')
    
    export_file = fields.Binary(string='Fichier exporté', readonly=True)
    export_filename = fields.Char(string='Nom du fichier', readonly=True)

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if self.env.context.get('active_ids'):
            res['article_ids'] = [(6, 0, self.env.context.get('active_ids'))]
        return res

    def action_export(self):
        """Lance l'export vers Naviwest"""
        self.ensure_one()
        
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("La bibliothèque openpyxl n'est pas installée."))
        
        if not self.article_ids:
            raise UserError(_('Veuillez sélectionner au moins un article à exporter.'))
        
        config = self.env['vst.config'].get_config()
        
        template_files = {
            'beg': config.template_beg,
            'niedax': config.template_niedax,
            'cables': config.template_cables,
        }
        
        template_path = template_files.get(self.export_type)
        
        if not template_path or not os.path.exists(template_path):
            type_names = {'beg': 'BEG', 'niedax': 'NIEDAX', 'cables': 'CÂBLES'}
            type_name = type_names.get(self.export_type)
            
            raise UserError(_(
                'Fichier template %s non trouvé.\n\n'
                'Chemin configuré : %s\n\n'
                'Vérifiez la configuration dans VST > Configuration > Configuration VST'
            ) % (type_name, template_path or '(non configuré)'))
        
        try:
            wb = load_workbook(template_path)
            
            if 'Article' in wb.sheetnames:
                ws = wb['Article']
            else:
                ws = wb.active
            
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            
            articles = self.article_ids
            if self.filter_field and self.filter_value:
                articles = articles.filtered(
                    lambda a: a[self.filter_field] and 
                    self.filter_value.upper() in str(a[self.filter_field]).upper()
                )
            
            if not articles:
                raise UserError(_('Aucun article correspondant aux filtres sélectionnés.'))
            
            current_row = 2
            
            for article in articles:
                ws.cell(current_row, 1, article.reference_fabricant or '')
                ws.cell(current_row, 2, article.designation or '')
                ws.cell(current_row, 3, (article.designation or '')[:50])
                ws.cell(current_row, 4, article.unite or 'U')
                ws.cell(current_row, 5, '30')  # Groupe compta
                ws.cell(current_row, 6, article.libelle_famille or '')
                ws.cell(current_row, 7, article.libelle_sous_famille or '')
                ws.cell(current_row, 8, article.prix_achat_adherent or 0)
                ws.cell(current_row, 9, article.prix_public_ht or 0)
                ws.cell(current_row, 10, 1)  # Quantité
                ws.cell(current_row, 11, article.nom_fabricant or '')
                
                current_row += 1
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            type_names = {'beg': 'BEG', 'niedax': 'NIEDAX', 'cables': 'CABLES'}
            filename = f"Export_Naviwest_{type_names[self.export_type]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            self.write({
                'export_file': base64.b64encode(output.read()),
                'export_filename': filename,
            })
            
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'export.naviwest.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }
            
        except Exception as e:
            _logger.error("Erreur export Naviwest: %s", str(e))
            raise UserError(_('Erreur lors de l\'export :\n%s') % str(e))

    def action_download(self):
        """Télécharge le fichier exporté"""
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model=export.naviwest.wizard&id={self.id}&field=export_file&filename_field=export_filename&download=true',
            'target': 'self',
        }
