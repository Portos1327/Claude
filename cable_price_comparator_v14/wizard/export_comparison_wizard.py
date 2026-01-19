# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import logging
from datetime import datetime

_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportComparisonWizard(models.TransientModel):
    """Wizard d'export des comparaisons de prix - Multi-formats"""
    _name = 'cable.export.comparison.wizard'
    _description = 'Export comparaison prix'

    export_format = fields.Selection([
        ('comparison', 'Tableau comparatif Excel'),
        ('naviwest', 'Format Naviwest'),
        ('quickdevis', 'Format QuickDevis 7'),
        ('report', 'Rapport statistique complet'),
    ], string='Format d\'export', default='comparison', required=True)
    
    export_scope = fields.Selection([
        ('all', 'Tous les produits maîtres'),
        ('selected', 'Produits sélectionnés'),
        ('multi_supplier', 'Multi-fournisseurs uniquement'),
        ('type', 'Par type de câble'),
    ], string='Portée', default='all', required=True)
    
    master_product_ids = fields.Many2many(
        'cable.product.master',
        string='Produits à exporter'
    )
    
    cable_type_ids = fields.Many2many(
        'cable.type',
        string='Types de câbles'
    )
    
    supplier_ids = fields.Many2many(
        'cable.supplier',
        string='Fournisseurs',
        help='Filtrer par fournisseurs (laisser vide pour tous)'
    )
    
    min_suppliers = fields.Integer(
        string='Minimum fournisseurs',
        default=2
    )
    
    include_details = fields.Boolean(
        string='Inclure détails références',
        default=True
    )
    
    include_variations = fields.Boolean(
        string='Inclure variations M-1',
        default=True
    )
    
    # Résultat
    file_data = fields.Binary(string='Fichier', readonly=True)
    file_name = fields.Char(string='Nom du fichier', readonly=True)
    state = fields.Selection([
        ('draft', 'Configuration'),
        ('done', 'Terminé'),
    ], default='draft')
    
    def action_export(self):
        """Générer l'export selon le format choisi"""
        self.ensure_one()
        
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("openpyxl n'est pas installé."))
        
        if self.export_format == 'comparison':
            return self._export_comparison()
        elif self.export_format == 'naviwest':
            return self._export_naviwest()
        elif self.export_format == 'quickdevis':
            return self._export_quickdevis()
        elif self.export_format == 'report':
            return self._export_report()
    
    def _get_products_to_export(self):
        """Récupère les produits maîtres selon la portée"""
        domain = []
        if self.export_scope == 'selected' and self.master_product_ids:
            domain.append(('id', 'in', self.master_product_ids.ids))
        elif self.export_scope == 'multi_supplier':
            domain.append(('supplier_count', '>=', self.min_suppliers))
        elif self.export_scope == 'type' and self.cable_type_ids:
            domain.append(('cable_type_id', 'in', self.cable_type_ids.ids))
        
        return self.env['cable.product.master'].search(domain, order='cable_type_code, section')
    
    def _get_suppliers(self, masters):
        """Récupère les fournisseurs pour les produits"""
        if self.supplier_ids:
            return self.supplier_ids.sorted(key=lambda s: s.name)
        
        supplier_ids = set()
        for master in masters:
            for line in master.pricelist_line_ids:
                supplier_ids.add(line.supplier_id.id)
        return self.env['cable.supplier'].browse(list(supplier_ids)).sorted(key=lambda s: s.name)
    
    def _get_freshness_color(self, freshness):
        """Retourne la couleur selon la fraîcheur du tarif"""
        colors = {
            'current': '92D050',   # Vert
            'previous': 'FFC000',  # Orange
            'old': 'FF6B6B',       # Rouge
        }
        return colors.get(freshness, 'FFFFFF')
    
    def _export_comparison(self):
        """Export tableau comparatif Excel avec code couleur"""
        masters = self._get_products_to_export()
        if not masters:
            raise UserError(_("Aucun produit à exporter."))
        
        suppliers = self._get_suppliers(masters)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparaison Prix ML"
        
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # En-têtes
        headers = ['Type', 'Config', 'Section', 'Désignation', 'Nb Fourn.']
        for supplier in suppliers:
            headers.append(f'{supplier.code} €/ml')
            headers.append(f'{supplier.code} Date')
            if self.include_variations:
                headers.append(f'{supplier.code} Var%')
            if self.include_details:
                headers.append(f'{supplier.code} Réf.')
        headers.extend(['Min €/ml', 'Max €/ml', 'Écart %', 'Meilleur', 'Var Moy%'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = PatternFill('solid', fgColor='4472C4')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        ws.row_dimensions[1].height = 35
        
        # Données
        row = 2
        for master in masters:
            ws.cell(row=row, column=1, value=master.cable_type_code or '').border = border
            ws.cell(row=row, column=2, value=master.conductor_config or '').border = border
            ws.cell(row=row, column=3, value=master.section or 0).border = border
            ws.cell(row=row, column=4, value=master.name[:50]).border = border
            ws.cell(row=row, column=5, value=master.supplier_count).border = border
            
            col = 6
            best_price = float('inf')
            best_col = None
            
            for supplier in suppliers:
                lines = master.pricelist_line_ids.filtered(
                    lambda l, s=supplier: l.supplier_id.id == s.id
                ).sorted(key=lambda l: l.date_tarif or fields.Date.today(), reverse=True)
                
                if lines:
                    line = lines[0]
                    price = line.price_per_ml
                    
                    price_cell = ws.cell(row=row, column=col, value=price)
                    price_cell.number_format = '0.0000'
                    price_cell.border = border
                    
                    freshness_color = self._get_freshness_color(line.tarif_freshness)
                    price_cell.fill = PatternFill('solid', fgColor=freshness_color)
                    
                    if price > 0 and price < best_price:
                        best_price = price
                        best_col = col
                    
                    col += 1
                    
                    date_cell = ws.cell(row=row, column=col, value=line.date_tarif)
                    date_cell.number_format = 'MM/YY'
                    date_cell.border = border
                    date_cell.fill = PatternFill('solid', fgColor=freshness_color)
                    col += 1
                    
                    if self.include_variations:
                        var_val = line.price_variation / 100 if line.price_variation else 0
                        var_cell = ws.cell(row=row, column=col, value=var_val)
                        var_cell.number_format = '+0.0%;-0.0%;0%'
                        var_cell.border = border
                        if line.price_trend == 'up':
                            var_cell.font = Font(color='FF0000', bold=True)
                        elif line.price_trend == 'down':
                            var_cell.font = Font(color='00B050', bold=True)
                        col += 1
                    
                    if self.include_details:
                        ws.cell(row=row, column=col, value=line.reference).border = border
                        col += 1
                else:
                    ws.cell(row=row, column=col, value='-').border = border
                    col += 1
                    ws.cell(row=row, column=col, value='-').border = border
                    col += 1
                    if self.include_variations:
                        ws.cell(row=row, column=col, value='-').border = border
                        col += 1
                    if self.include_details:
                        ws.cell(row=row, column=col, value='-').border = border
                        col += 1
            
            # Stats
            if best_price < float('inf'):
                ws.cell(row=row, column=col, value=master.price_min).number_format = '0.0000'
                ws.cell(row=row, column=col).border = border
                col += 1
                ws.cell(row=row, column=col, value=master.price_max).number_format = '0.0000'
                ws.cell(row=row, column=col).border = border
                col += 1
                ws.cell(row=row, column=col, value=master.price_spread / 100).number_format = '0.0%'
                ws.cell(row=row, column=col).border = border
                col += 1
                ws.cell(row=row, column=col, value=master.best_supplier_id.code if master.best_supplier_id else '').border = border
                col += 1
                ws.cell(row=row, column=col, value=(master.avg_price_variation or 0) / 100).number_format = '+0.0%;-0.0%;0%'
                ws.cell(row=row, column=col).border = border
                
                if best_col:
                    ws.cell(row=row, column=best_col).fill = PatternFill('solid', fgColor='C6EFCE')
            
            row += 1
        
        # Légende
        row += 2
        ws.cell(row=row, column=1, value="LÉGENDE:").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Vert").fill = PatternFill('solid', fgColor='92D050')
        ws.cell(row=row, column=2, value="= Mois courant")
        row += 1
        ws.cell(row=row, column=1, value="Orange").fill = PatternFill('solid', fgColor='FFC000')
        ws.cell(row=row, column=2, value="= Mois précédent")
        row += 1
        ws.cell(row=row, column=1, value="Rouge").fill = PatternFill('solid', fgColor='FF6B6B')
        ws.cell(row=row, column=2, value="= Ancien (> 2 mois)")
        row += 1
        ws.cell(row=row, column=1, value="Vert foncé").fill = PatternFill('solid', fgColor='C6EFCE')
        ws.cell(row=row, column=2, value="= Meilleur prix")
        
        # Ajuster largeurs
        for i, width in enumerate([8, 10, 7, 35, 6], 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        self.file_data = base64.b64encode(output.read())
        self.file_name = f"comparaison_prix_ml_{fields.Date.today()}.xlsx"
        self.state = 'done'
        
        return self._return_wizard()
    
    def _export_naviwest(self):
        """Export au format Naviwest"""
        masters = self._get_products_to_export()
        if not masters:
            raise UserError(_("Aucun produit à exporter."))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Export Naviwest"
        
        headers = ['CODE_ARTICLE', 'DESIGNATION', 'UNITE', 'PRIX_ACHAT_HT', 
                   'FOURNISSEUR', 'REF_FOURNISSEUR', 'FAMILLE', 'SECTION', 
                   'NB_CONDUCTEURS', 'DATE_TARIF']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        row = 2
        for master in masters:
            lines = master.pricelist_line_ids
            if self.supplier_ids:
                lines = lines.filtered(lambda l: l.supplier_id in self.supplier_ids)
            
            suppliers_done = set()
            for line in lines.sorted(key=lambda l: l.date_tarif or fields.Date.today(), reverse=True):
                if line.supplier_id.id in suppliers_done:
                    continue
                suppliers_done.add(line.supplier_id.id)
                
                ws.cell(row=row, column=1, value=master.matching_key or str(master.id))
                ws.cell(row=row, column=2, value=master.name[:60])
                ws.cell(row=row, column=3, value='ML')
                ws.cell(row=row, column=4, value=round(line.price_per_ml, 4))
                ws.cell(row=row, column=5, value=line.supplier_code)
                ws.cell(row=row, column=6, value=line.reference)
                ws.cell(row=row, column=7, value=master.cable_type_code)
                ws.cell(row=row, column=8, value=master.section)
                ws.cell(row=row, column=9, value=master.nb_conductors)
                ws.cell(row=row, column=10, value=line.date_tarif)
                row += 1
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        self.file_data = base64.b64encode(output.read())
        self.file_name = f"export_naviwest_{fields.Date.today()}.xlsx"
        self.state = 'done'
        
        return self._return_wizard()
    
    def _export_quickdevis(self):
        """Export au format QuickDevis 7"""
        masters = self._get_products_to_export()
        if not masters:
            raise UserError(_("Aucun produit à exporter."))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Export QuickDevis"
        
        headers = ['Reference', 'Designation', 'Unite', 'PrixUnitaire', 
                   'CodeFournisseur', 'RefFournisseur', 'Famille1', 'Famille2',
                   'Coef', 'Remise', 'DateMAJ']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        row = 2
        for master in masters:
            best_line = None
            best_price = float('inf')
            
            for line in master.pricelist_line_ids:
                if line.price_per_ml > 0 and line.price_per_ml < best_price:
                    best_price = line.price_per_ml
                    best_line = line
            
            if not best_line:
                continue
            
            ref_qdv = f"CAB-{master.cable_type_code or 'XX'}-{master.conductor_config or ''}"
            ref_qdv = ref_qdv.replace(',', '').replace(' ', '')[:20]
            
            ws.cell(row=row, column=1, value=ref_qdv)
            ws.cell(row=row, column=2, value=master.name[:80])
            ws.cell(row=row, column=3, value='ML')
            ws.cell(row=row, column=4, value=round(best_line.price_per_ml, 4))
            ws.cell(row=row, column=5, value=best_line.supplier_code)
            ws.cell(row=row, column=6, value=best_line.reference)
            ws.cell(row=row, column=7, value=master.cable_type_code or 'CABLE')
            ws.cell(row=row, column=8, value=f"{master.section or 0} mm²")
            ws.cell(row=row, column=9, value=1)
            ws.cell(row=row, column=10, value=0)
            ws.cell(row=row, column=11, value=best_line.date_tarif or fields.Date.today())
            row += 1
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        self.file_data = base64.b64encode(output.read())
        self.file_name = f"export_quickdevis_{fields.Date.today()}.xlsx"
        self.state = 'done'
        
        return self._return_wizard()
    
    def _export_report(self):
        """Export rapport statistique complet"""
        masters = self._get_products_to_export()
        if not masters:
            raise UserError(_("Aucun produit à exporter."))
        
        suppliers = self._get_suppliers(masters)
        
        wb = Workbook()
        
        # FEUILLE 1: SYNTHÈSE
        ws = wb.active
        ws.title = "Synthèse"
        
        ws.cell(row=1, column=1, value="RAPPORT COMPARATIF PRIX CÂBLES").font = Font(bold=True, size=16)
        ws.cell(row=2, column=1, value=f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        row = 4
        ws.cell(row=row, column=1, value="STATISTIQUES GÉNÉRALES").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value="Produits maîtres:")
        ws.cell(row=row, column=2, value=len(masters))
        row += 1
        ws.cell(row=row, column=1, value="Fournisseurs:")
        ws.cell(row=row, column=2, value=len(suppliers))
        
        row += 2
        ws.cell(row=row, column=1, value="PAR FOURNISSEUR").font = Font(bold=True, size=12)
        row += 1
        for col, h in enumerate(['Fournisseur', 'Articles', 'Meilleur sur', '% Meilleur'], 1):
            ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        row += 1
        
        for supplier in suppliers:
            line_count = sum(1 for m in masters for l in m.pricelist_line_ids if l.supplier_id == supplier)
            best_count = len(masters.filtered(lambda m: m.best_supplier_id == supplier))
            pct = (best_count / len(masters) * 100) if masters else 0
            
            ws.cell(row=row, column=1, value=supplier.name)
            ws.cell(row=row, column=2, value=line_count)
            ws.cell(row=row, column=3, value=best_count)
            ws.cell(row=row, column=4, value=f"{pct:.1f}%")
            row += 1
        
        row += 2
        ws.cell(row=row, column=1, value="PAR TYPE DE CÂBLE").font = Font(bold=True, size=12)
        row += 1
        for col, h in enumerate(['Type', 'Nb', 'Min €/ml', 'Max €/ml', 'Moy €/ml', 'Écart moy%'], 1):
            ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        row += 1
        
        cable_types = {}
        for master in masters:
            t = master.cable_type_code or 'AUTRE'
            cable_types.setdefault(t, []).append(master)
        
        for type_code, type_masters in sorted(cable_types.items()):
            prices = [m.price_avg for m in type_masters if m.price_avg > 0]
            spreads = [m.price_spread for m in type_masters if m.price_spread]
            
            ws.cell(row=row, column=1, value=type_code)
            ws.cell(row=row, column=2, value=len(type_masters))
            ws.cell(row=row, column=3, value=round(min(prices), 4) if prices else 0)
            ws.cell(row=row, column=4, value=round(max(prices), 4) if prices else 0)
            ws.cell(row=row, column=5, value=round(sum(prices)/len(prices), 4) if prices else 0)
            ws.cell(row=row, column=6, value=f"{sum(spreads)/len(spreads):.1f}%" if spreads else "0%")
            row += 1
        
        row += 2
        ws.cell(row=row, column=1, value="ÉVOLUTION M-1").font = Font(bold=True, size=12)
        row += 1
        
        trends = {'up': 0, 'down': 0, 'stable': 0, 'new': 0}
        for m in masters:
            for l in m.pricelist_line_ids:
                if l.price_trend:
                    trends[l.price_trend] = trends.get(l.price_trend, 0) + 1
        
        ws.cell(row=row, column=1, value="↗ Hausse:")
        ws.cell(row=row, column=2, value=trends['up']).font = Font(color='FF0000')
        row += 1
        ws.cell(row=row, column=1, value="↘ Baisse:")
        ws.cell(row=row, column=2, value=trends['down']).font = Font(color='00B050')
        row += 1
        ws.cell(row=row, column=1, value="→ Stables:")
        ws.cell(row=row, column=2, value=trends['stable'])
        row += 1
        ws.cell(row=row, column=1, value="✦ Nouveaux:")
        ws.cell(row=row, column=2, value=trends['new'])
        
        # FEUILLE 2: TOP ÉCARTS
        ws2 = wb.create_sheet("Top Écarts")
        headers = ['Type', 'Config', 'Désignation', 'Min €/ml', 'Fourn Min', 'Max €/ml', 'Fourn Max', 'Écart %', 'Éco/1000m']
        for col, h in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=h).font = Font(bold=True)
        
        row = 2
        for master in sorted(masters, key=lambda m: m.price_spread or 0, reverse=True)[:50]:
            if master.price_spread < 5:
                continue
            max_line = max(master.pricelist_line_ids, key=lambda l: l.price_per_ml, default=None)
            
            ws2.cell(row=row, column=1, value=master.cable_type_code)
            ws2.cell(row=row, column=2, value=master.conductor_config)
            ws2.cell(row=row, column=3, value=master.name[:40])
            ws2.cell(row=row, column=4, value=round(master.price_min, 4))
            ws2.cell(row=row, column=5, value=master.best_supplier_id.code if master.best_supplier_id else '')
            ws2.cell(row=row, column=6, value=round(master.price_max, 4))
            ws2.cell(row=row, column=7, value=max_line.supplier_code if max_line else '')
            ws2.cell(row=row, column=8, value=f"{master.price_spread:.1f}%")
            eco = (master.price_max - master.price_min) * 1000 if master.price_max and master.price_min else 0
            ws2.cell(row=row, column=9, value=round(eco, 2)).number_format = '#,##0.00 €'
            row += 1
        
        # FEUILLE 3: DÉTAIL
        ws3 = wb.create_sheet("Détail")
        headers = ['Type', 'Config', 'Section', 'Désignation', 'Fournisseur', 'Référence', 
                   'Prix €/ml', 'Date', 'Fraîcheur', 'Var %', 'Tendance']
        for col, h in enumerate(headers, 1):
            ws3.cell(row=1, column=col, value=h).font = Font(bold=True)
        
        row = 2
        for master in masters:
            for line in master.pricelist_line_ids.sorted(key=lambda l: l.supplier_code or ''):
                ws3.cell(row=row, column=1, value=master.cable_type_code)
                ws3.cell(row=row, column=2, value=master.conductor_config)
                ws3.cell(row=row, column=3, value=master.section)
                ws3.cell(row=row, column=4, value=master.name[:40])
                ws3.cell(row=row, column=5, value=line.supplier_code)
                ws3.cell(row=row, column=6, value=line.reference)
                price_cell = ws3.cell(row=row, column=7, value=round(line.price_per_ml, 4))
                color = self._get_freshness_color(line.tarif_freshness)
                price_cell.fill = PatternFill('solid', fgColor=color)
                ws3.cell(row=row, column=8, value=line.date_tarif)
                
                freshness_labels = {'current': 'Actuel', 'previous': 'M-1', 'old': 'Ancien'}
                ws3.cell(row=row, column=9, value=freshness_labels.get(line.tarif_freshness, ''))
                ws3.cell(row=row, column=10, value=f"{line.price_variation:.1f}%" if line.price_variation else '')
                
                trend_labels = {'up': '↗ Hausse', 'down': '↘ Baisse', 'stable': '→ Stable', 'new': '✦ Nouveau'}
                ws3.cell(row=row, column=11, value=trend_labels.get(line.price_trend, ''))
                row += 1
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        self.file_data = base64.b64encode(output.read())
        self.file_name = f"rapport_comparatif_{fields.Date.today()}.xlsx"
        self.state = 'done'
        
        return self._return_wizard()
    
    def _return_wizard(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
