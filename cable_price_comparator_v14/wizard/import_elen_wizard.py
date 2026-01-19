# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import tempfile
import re
import logging

_logger = logging.getLogger(__name__)


class ImportElenWizard(models.TransientModel):
    """Wizard d'import tarif ELEN Distribution (Excel XLSX)
    
    Structure du fichier ELEN:
    - Ligne 1: Dates (début et fin validité)
    - Ligne 2: "TARIF CÂBLE"
    - Ligne 3: N° tarif
    - Ligne 5: En-têtes (Famille, Sous-Famille, Code Article, Codet EDF, Libellé Article, € / KM)
    - Ligne 6+: Données
    
    IMPORTANT: Prix en €/km dans le fichier, convertis en €/ml à l'import
    """
    _name = 'cable.import.elen.wizard'
    _description = 'Import Tarif ELEN'

    file_data = fields.Binary(
        string='Fichier Excel ELEN',
        required=True,
        help='Fichier Excel ELEN Distribution (.xlsx)'
    )
    file_name = fields.Char(string='Nom du fichier')
    
    supplier_id = fields.Many2one(
        'cable.supplier',
        string='Fournisseur',
        required=True,
        default=lambda self: self._default_supplier()
    )
    
    date_validity = fields.Date(
        string='Date de validité',
        default=fields.Date.today,
        required=True
    )
    
    # Options
    create_master_products = fields.Boolean(
        string='Créer les produits maîtres',
        default=True,
        help='Créer automatiquement les produits maîtres manquants'
    )
    update_existing = fields.Boolean(
        string='Mettre à jour les prix existants',
        default=True
    )
    
    # Stats
    lines_processed = fields.Integer(string='Lignes traitées', readonly=True)
    lines_created = fields.Integer(string='Lignes créées', readonly=True)
    lines_updated = fields.Integer(string='Lignes mises à jour', readonly=True)
    masters_created = fields.Integer(string='Produits maîtres créés', readonly=True)
    errors = fields.Text(string='Erreurs', readonly=True)
    
    def _default_supplier(self):
        """Cherche ou crée le fournisseur ELEN"""
        supplier = self.env['cable.supplier'].search([
            ('code', '=', 'ELEN')
        ], limit=1)
        if not supplier:
            supplier = self.env['cable.supplier'].create({
                'name': 'ELEN Distribution',
                'code': 'ELEN',
                'supplier_type': 'distributor',
                'default_price_unit': 'km',
            })
        return supplier
    
    def action_import(self):
        """Importer le fichier ELEN"""
        self.ensure_one()
        
        if not self.file_data:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))
        
        try:
            import openpyxl
        except ImportError:
            raise UserError(_("La bibliothèque openpyxl n'est pas installée."))
        
        # Sauvegarder le fichier temporairement
        file_content = base64.b64decode(self.file_data)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(file_content)
            tmp_path = tmp.name
        
        # Ouvrir le fichier Excel
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        ws = wb.active
        
        # Créer ou récupérer le tarif
        pricelist = self._get_or_create_pricelist()
        
        # Stats
        stats = {
            'processed': 0,
            'created': 0,
            'updated': 0,
            'masters_created': 0,
            'errors': []
        }
        
        # Lire la famille courante (car elle n'est pas répétée sur chaque ligne)
        current_famille = ""
        current_sous_famille = ""
        
        # Parcourir les lignes (à partir de la ligne 6)
        for row_idx in range(6, ws.max_row + 1):
            row = ws[row_idx]
            
            # Récupérer les valeurs
            famille = row[0].value
            sous_famille = row[1].value
            code_article = row[2].value
            codet_edf = row[3].value
            libelle = row[4].value
            prix_km = row[5].value
            
            # Mettre à jour famille si présente
            if famille:
                current_famille = str(famille).strip()
            if sous_famille:
                current_sous_famille = str(sous_famille).strip()
            
            # Ignorer si pas de code article ou libellé
            if not code_article or not libelle:
                continue
            
            stats['processed'] += 1
            
            try:
                # Parser le libellé pour extraire les caractéristiques
                type_code, nb_cond, has_ground, section = self._parse_libelle(str(libelle))
                
                if not type_code:
                    stats['errors'].append(f"L{row_idx}: Type non reconnu - {libelle}")
                    continue
                
                # Convertir prix €/km en €/ml
                prix_ml = 0
                if prix_km:
                    try:
                        prix_ml = float(prix_km) / 1000.0
                    except:
                        stats['errors'].append(f"L{row_idx}: Prix invalide - {prix_km}")
                        continue
                
                # Chercher ou créer la ligne
                line = self._find_or_create_line(
                    pricelist=pricelist,
                    reference=str(code_article).strip(),
                    designation=str(libelle).strip(),
                    type_code=type_code,
                    nb_conductors=nb_cond,
                    has_ground=has_ground,
                    section=section,
                    prix_ml=prix_ml,
                    famille=current_famille,
                    ean=str(codet_edf).strip() if codet_edf else '',
                    stats=stats
                )
                
                # Créer produit maître si demandé
                if self.create_master_products and line and not line.master_product_id:
                    line.action_find_or_create_master()
                    if line.master_product_id:
                        stats['masters_created'] += 1
                        
            except Exception as e:
                stats['errors'].append(f"L{row_idx}: {str(e)}")
                _logger.warning(f"Erreur import ELEN ligne {row_idx}: {e}")
        
        # Fermer le workbook AVANT de supprimer le fichier (obligatoire sur Windows)
        wb.close()
        
        # Nettoyer le fichier temporaire
        import os
        try:
            os.unlink(tmp_path)
        except Exception as e:
            _logger.warning(f"Impossible de supprimer le fichier temporaire: {e}")
        
        # Mettre à jour les stats
        self.write({
            'lines_processed': stats['processed'],
            'lines_created': stats['created'],
            'lines_updated': stats['updated'],
            'masters_created': stats['masters_created'],
            'errors': '\n'.join(stats['errors'][:50]) if stats['errors'] else '',
        })
        
        # Message de résultat
        message = _(
            "Import ELEN terminé:\n"
            "- Lignes traitées: %(processed)d\n"
            "- Lignes créées: %(created)d\n"
            "- Lignes mises à jour: %(updated)d\n"
            "- Produits maîtres créés: %(masters_created)d\n"
            "- Erreurs: %(errors)d"
        ) % {
            'processed': stats['processed'],
            'created': stats['created'],
            'updated': stats['updated'],
            'masters_created': stats['masters_created'],
            'errors': len(stats['errors']),
        }
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import ELEN'),
                'message': message,
                'type': 'success' if not stats['errors'] else 'warning',
                'sticky': True,
            }
        }
    
    def _get_or_create_pricelist(self):
        """Crée ou récupère le tarif ELEN"""
        # Chercher un tarif existant pour ce mois
        existing = self.env['cable.pricelist'].search([
            ('supplier_id', '=', self.supplier_id.id),
            ('date_validity', '=', self.date_validity),
        ], limit=1)
        
        if existing and not self.update_existing:
            raise UserError(_("Un tarif existe déjà pour cette date. Cochez 'Mettre à jour' pour le modifier."))
        
        if existing:
            return existing
        
        return self.env['cable.pricelist'].create({
            'name': f"ELEN {self.date_validity.strftime('%Y-%m')}",
            'supplier_id': self.supplier_id.id,
            'date_validity': self.date_validity,
            'date_import': fields.Datetime.now(),
            'state': 'imported',
        })
    
    def _parse_libelle(self, libelle):
        """Parse le libellé ELEN pour extraire type, config, section
        
        Exemples:
        - "AR2V 1 X 50 NORME NF C32-321 : 2022" -> AR2V, 1, False, 50
        - "R2V 3 G 1,5 NORME NF C32-321 : 2022" -> R2V, 3, True, 1.5
        - "R2V 5 G 2,5 NORME NF C32-321 : 2022" -> R2V, 5, True, 2.5
        
        Returns:
            tuple (type_code, nb_conductors, has_ground, section)
        """
        if not libelle:
            return ('', 1, True, 0)
        
        libelle = libelle.upper()
        
        # Extraire le type
        type_code = ''
        type_patterns = [
            r'^(AR2V)',
            r'^(R2V)',
            r'^(H07V-[URK])',
            r'^(H07RN-F)',
            r'^(U-?1000)',
        ]
        for pattern in type_patterns:
            match = re.search(pattern, libelle)
            if match:
                type_code = match.group(1).replace(' ', '')
                break
        
        if not type_code:
            return ('', 1, True, 0)
        
        # Pattern pour extraire nb + G/X + section
        # Ex: "3 G 1,5" ou "1 X 50" ou "4 X 16"
        match = re.search(r'(\d+)\s*([GX])\s*([\d,\.]+)', libelle)
        if match:
            nb = int(match.group(1))
            has_ground = match.group(2) == 'G'
            section = float(match.group(3).replace(',', '.'))
            return (type_code, nb, has_ground, section)
        
        return (type_code, 1, True, 0)
    
    def _find_or_create_line(self, pricelist, reference, designation, type_code, 
                             nb_conductors, has_ground, section, prix_ml, famille, ean, stats):
        """Trouve ou crée une ligne de tarif"""
        
        Line = self.env['cable.pricelist.line']
        
        # Chercher ligne existante
        existing = Line.search([
            ('pricelist_id', '=', pricelist.id),
            ('reference', '=', reference),
        ], limit=1)
        
        vals = {
            'designation': designation,
            'cable_type_code': type_code,
            'nb_conductors': nb_conductors,
            'has_ground': has_ground,
            'section': section,
            'price_net': prix_ml,
            'price_unit': 'm',  # Prix déjà converti en €/ml
            'family': famille,
            'ean': ean,
        }
        
        if existing:
            if self.update_existing:
                # Sauvegarder l'ancien prix comme prix M-1
                if existing.price_per_ml and existing.price_per_ml != prix_ml:
                    vals['price_previous_month'] = existing.price_per_ml
                existing.write(vals)
                stats['updated'] += 1
            return existing
        else:
            vals.update({
                'pricelist_id': pricelist.id,
                'reference': reference,
            })
            line = Line.create(vals)
            stats['created'] += 1
            return line
