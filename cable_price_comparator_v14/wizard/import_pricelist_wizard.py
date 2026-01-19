# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import base64
import io
import logging

_logger = logging.getLogger(__name__)

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    _logger.warning("pandas non disponible - import Excel limité")

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ImportPricelistWizard(models.TransientModel):
    """Wizard d'import de tarifs câbles depuis Excel"""
    _name = 'cable.import.pricelist.wizard'
    _description = 'Import tarif câbles'

    # Source
    import_mode = fields.Selection([
        ('excel', 'Fichier Excel'),
        ('rexel', 'Module Rexel (articles câbles)'),
    ], string='Mode d\'import', default='excel', required=True)
    
    # Fichier Excel
    file_data = fields.Binary(string='Fichier Excel', attachment=False)
    file_name = fields.Char(string='Nom du fichier')
    
    # Fournisseur
    supplier_id = fields.Many2one(
        'cable.supplier',
        string='Fournisseur',
        required=True
    )
    
    # Configuration
    config_id = fields.Many2one(
        'cable.supplier.excel.config',
        string='Configuration Excel',
        domain="[('supplier_id', '=', supplier_id)]"
    )
    
    # Sélection de feuille
    sheet_names = fields.Char(string='Feuilles disponibles', readonly=True)
    sheet_name = fields.Char(
        string='Feuille à importer',
        help='Nom de la feuille Excel (laisser vide pour la première)'
    )
    
    # Mapping colonnes
    col_reference = fields.Char(string='Colonne Référence', default='REFARTICLE')
    col_designation = fields.Char(string='Colonne Désignation', default='Désignation')
    col_price_gross = fields.Char(string='Colonne Prix Brut', default='BRUT')
    col_discount = fields.Char(string='Colonne Remise', default='Remise')
    col_price_net = fields.Char(string='Colonne Prix Net', default='NET')
    col_family = fields.Char(string='Colonne Famille', default='Famille')
    col_unit = fields.Char(string='Colonne Unité')
    col_weight = fields.Char(string='Colonne Poids', default='POIDS')
    col_ean = fields.Char(string='Colonne EAN')
    col_datasheet_url = fields.Char(string='Colonne Fiche technique', default='Lien fiche technique')
    
    header_row = fields.Integer(string='Ligne d\'en-tête', default=0)
    
    # Unité de prix
    price_unit = fields.Selection([
        ('m', '€/m'),
        ('km', '€/km'),
        ('100m', '€/100m'),
        ('unit', '€/unité'),
        ('kg', '€/kg'),
    ], string='Unité de prix', default='m', required=True,
       help='Rexel et la plupart des fournisseurs utilisent €/ml')
    
    # Dates
    date_validity = fields.Date(
        string='Date de validité',
        default=fields.Date.today
    )
    period_name = fields.Char(string='Période tarifaire')
    
    # Options
    extract_characteristics = fields.Boolean(
        string='Extraire caractéristiques',
        default=True,
        help='Extraire type câble, section, etc. des désignations'
    )
    run_matching = fields.Boolean(
        string='Lancer le matching automatique',
        default=True,
        help='Chercher les correspondances avec les produits maîtres'
    )
    create_masters = fields.Boolean(
        string='Créer produits maîtres',
        default=True,
        help='Créer automatiquement les produits maîtres non existants'
    )
    
    # Résultats
    result_log = fields.Text(string='Log d\'import', readonly=True)
    pricelist_id = fields.Many2one(
        'cable.pricelist',
        string='Tarif créé',
        readonly=True
    )
    
    state = fields.Selection([
        ('draft', 'Configuration'),
        ('preview', 'Aperçu'),
        ('done', 'Terminé'),
    ], string='État', default='draft')
    
    preview_html = fields.Html(string='Aperçu', readonly=True)
    
    @api.onchange('supplier_id')
    def _onchange_supplier_id(self):
        """Charger les valeurs par défaut du fournisseur"""
        if self.supplier_id:
            self.col_reference = self.supplier_id.default_ref_column or 'REFARTICLE'
            self.col_designation = self.supplier_id.default_designation_column or 'Désignation'
            self.col_price_net = self.supplier_id.default_price_column or 'NET'
            self.col_family = self.supplier_id.default_family_column or 'Famille'
            self.price_unit = self.supplier_id.default_price_unit or 'km'
            
            config = self.env['cable.supplier.excel.config'].search([
                ('supplier_id', '=', self.supplier_id.id),
                ('active', '=', True)
            ], limit=1)
            if config:
                self.config_id = config.id
    
    @api.onchange('config_id')
    def _onchange_config_id(self):
        """Appliquer la configuration sélectionnée"""
        if self.config_id:
            config = self.config_id
            self.sheet_name = config.sheet_name
            self.header_row = config.header_row
            self.col_reference = config.col_reference
            self.col_designation = config.col_designation
            self.col_price_gross = config.col_price_gross
            self.col_discount = config.col_discount
            self.col_price_net = config.col_price_net
            self.col_family = config.col_family
            self.col_unit = config.col_unit
            self.col_weight = config.col_weight
            self.col_ean = config.col_ean
            self.col_datasheet_url = config.col_datasheet_url
            self.price_unit = config.price_unit
    
    @api.onchange('file_data')
    def _onchange_file_data(self):
        """Analyser le fichier et lister les feuilles"""
        if not self.file_data:
            self.sheet_names = ''
            return
        
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("openpyxl n'est pas installé."))
        
        try:
            file_content = base64.b64decode(self.file_data)
            wb = load_workbook(io.BytesIO(file_content), read_only=True)
            self.sheet_names = ', '.join(wb.sheetnames)
            wb.close()
        except Exception as e:
            self.sheet_names = f'Erreur: {e}'
    
    def action_preview(self):
        """Afficher un aperçu des données"""
        self.ensure_one()
        
        if not self.file_data:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))
        
        if not PANDAS_AVAILABLE:
            raise UserError(_("pandas n'est pas installé."))
        
        try:
            file_content = base64.b64decode(self.file_data)
            
            df = pd.read_excel(
                io.BytesIO(file_content),
                sheet_name=self.sheet_name or 0,
                header=self.header_row
            )
            
            html = '<div class="table-responsive"><table class="table table-sm table-striped">'
            html += '<thead><tr>'
            for col in df.columns[:10]:
                html += f'<th>{col}</th>'
            html += '</tr></thead><tbody>'
            
            for idx, row in df.head(10).iterrows():
                html += '<tr>'
                for col in df.columns[:10]:
                    val = row[col] if pd.notna(row[col]) else ''
                    html += f'<td>{val}</td>'
                html += '</tr>'
            
            html += '</tbody></table></div>'
            html += f'<p class="text-muted">Total: {len(df)} lignes, {len(df.columns)} colonnes</p>'
            
            self.preview_html = html
            self.state = 'preview'
            
        except Exception as e:
            raise UserError(_(f"Erreur lecture fichier: {e}"))
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
    
    def action_import(self):
        """Lancer l'import"""
        self.ensure_one()
        
        if self.import_mode == 'excel':
            return self._import_from_excel()
        elif self.import_mode == 'rexel':
            return self._import_from_rexel()
    
    def _import_from_excel(self):
        """Import depuis fichier Excel"""
        if not self.file_data:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))
        
        if not PANDAS_AVAILABLE:
            raise UserError(_("pandas n'est pas installé."))
        
        try:
            file_content = base64.b64decode(self.file_data)
            
            df = pd.read_excel(
                io.BytesIO(file_content),
                sheet_name=self.sheet_name or 0,
                header=self.header_row
            )
            
            _logger.info(f"Import Excel: {len(df)} lignes, colonnes: {list(df.columns)}")
            
            # Créer le tarif
            pricelist = self.env['cable.pricelist'].create({
                'supplier_id': self.supplier_id.id,
                'date_validity': self.date_validity,
                'period_name': self.period_name,
                'file_name': self.file_name,
                'file_data': self.file_data,
                'sheet_name': self.sheet_name,
                'state': 'draft',
            })
            
            lines_data = []
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    ref = self._get_cell_value(row, self.col_reference)
                    if not ref:
                        continue
                    
                    designation = self._get_cell_value(row, self.col_designation)
                    if not designation:
                        continue
                    
                    price_gross = self._get_numeric_value(row, self.col_price_gross)
                    discount = self._get_numeric_value(row, self.col_discount)
                    price_net = self._get_numeric_value(row, self.col_price_net)
                    
                    if not price_net and price_gross:
                        if discount:
                            price_net = price_gross * (1 - discount / 100)
                        else:
                            price_net = price_gross
                    
                    line_vals = {
                        'pricelist_id': pricelist.id,
                        'reference': str(ref),
                        'designation': str(designation),
                        'price_gross': price_gross,
                        'discount': discount,
                        'price_net': price_net,
                        'price_unit': self.price_unit,
                        'family': self._get_cell_value(row, self.col_family),
                        'weight': self._get_numeric_value(row, self.col_weight),
                        'ean': self._get_cell_value(row, self.col_ean),
                        'datasheet_url': self._get_cell_value(row, self.col_datasheet_url),
                    }
                    
                    lines_data.append(line_vals)
                    
                except Exception as e:
                    errors.append(f"Ligne {idx + 1}: {e}")
            
            if lines_data:
                self.env['cable.pricelist.line'].create(lines_data)
            
            pricelist.state = 'imported'
            
            if self.extract_characteristics:
                _logger.info("Extraction des caractéristiques...")
                pricelist.line_ids.action_extract_characteristics()
            
            matched_count = 0
            created_count = 0
            if self.run_matching:
                _logger.info("Lancement du matching...")
                engine = self.env['cable.matching.engine']
                stats = engine.run_matching_batch(
                    pricelist_ids=[pricelist.id],
                    create_masters=self.create_masters
                )
                matched_count = stats.get('matched', 0)
                created_count = stats.get('created', 0)
                pricelist.state = 'matched'
            
            log_lines = [
                f"✅ Import terminé avec succès",
                f"",
                f"📊 Statistiques:",
                f"  • Lignes importées: {len(lines_data)}",
                f"  • Correspondances trouvées: {matched_count}",
                f"  • Produits maîtres créés: {created_count}",
                f"  • Erreurs: {len(errors)}",
            ]
            
            if errors[:10]:
                log_lines.append(f"")
                log_lines.append(f"⚠️ Premières erreurs:")
                for err in errors[:10]:
                    log_lines.append(f"  • {err}")
            
            self.result_log = '\n'.join(log_lines)
            self.pricelist_id = pricelist.id
            self.state = 'done'
            
            return {
                'type': 'ir.actions.act_window',
                'res_model': self._name,
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }
            
        except Exception as e:
            _logger.error(f"Erreur import: {e}")
            raise UserError(_(f"Erreur lors de l'import: {e}"))
    
    def _import_from_rexel(self):
        """Import depuis le module Rexel Article Manager"""
        try:
            RexelArticle = self.env['rexel.article']
        except KeyError:
            raise UserError(_(
                "Le module Rexel Article Manager n'est pas installé.\n"
                "Veuillez l'installer ou utiliser l'import Excel."
            ))
        
        cable_families = ['Câbles', 'CABLES', 'câble', 'cable', 'Fil', 'FIL']
        
        domain = ['|'] * (len(cable_families) - 1)
        for family in cable_families:
            domain.append(('famille_libelle', 'ilike', family))
        
        articles = RexelArticle.search(domain)
        
        if not articles:
            raise UserError(_("Aucun article câble trouvé dans le module Rexel."))
        
        pricelist = self.env['cable.pricelist'].create({
            'supplier_id': self.supplier_id.id,
            'date_validity': self.date_validity,
            'period_name': self.period_name or 'Import Rexel',
            'state': 'draft',
            'notes': f'Import depuis Rexel Article Manager - {len(articles)} articles',
        })
        
        lines_data = []
        for article in articles:
            # Récupérer les infos fabricant et distributeur
            line_vals = {
                'pricelist_id': pricelist.id,
                'reference': article.reference_fabricant or article.reference_rexel,
                'designation': article.designation,
                'price_gross': article.prix_base or 0,
                'discount': article.remise or 0,
                'price_net': article.prix_net or 0,
                'price_unit': 'm',  # Rexel = €/ml (mètre linéaire)
                'family': article.famille_libelle,
                'ean': article.code_ean13,
                # Fabricant
                'manufacturer_ref': article.reference_fabricant if hasattr(article, 'reference_fabricant') else '',
                'manufacturer_name': article.marque if hasattr(article, 'marque') else '',
                # Distributeur (Rexel)
                'distributor_ref': article.reference_rexel if hasattr(article, 'reference_rexel') else '',
                'distributor_name': 'Rexel',
                # Documents et images
                'image_url': article.url_image if hasattr(article, 'url_image') else '',
                'datasheet_url': article.url_fiche_technique if hasattr(article, 'url_fiche_technique') else '',
            }
            lines_data.append(line_vals)
        
        if lines_data:
            self.env['cable.pricelist.line'].create(lines_data)
        
        pricelist.state = 'imported'
        
        if self.extract_characteristics:
            pricelist.line_ids.action_extract_characteristics()
        
        if self.run_matching:
            engine = self.env['cable.matching.engine']
            engine.run_matching_batch(
                pricelist_ids=[pricelist.id],
                create_masters=self.create_masters
            )
            pricelist.state = 'matched'
        
        self.result_log = f"✅ Import Rexel terminé\n\n• {len(lines_data)} articles importés"
        self.pricelist_id = pricelist.id
        self.state = 'done'
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
    
    def _get_cell_value(self, row, column_name):
        """Récupère une valeur de cellule de manière sécurisée"""
        if not column_name:
            return None
        
        col_name_clean = column_name.strip().lower()
        
        for col in row.index:
            if str(col).strip().lower() == col_name_clean:
                val = row[col]
                if pd.isna(val):
                    return None
                return str(val).strip() if val else None
        
        return None
    
    def _get_numeric_value(self, row, column_name):
        """Récupère une valeur numérique"""
        val = self._get_cell_value(row, column_name)
        if not val:
            return 0.0
        
        val = str(val).replace(',', '.').replace(' ', '').replace('€', '')
        val = val.replace('Net', '').replace('net', '').strip()
        
        if not val or val == '-':
            return 0.0
        
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0
    
    def action_open_pricelist(self):
        """Ouvrir le tarif créé"""
        self.ensure_one()
        if not self.pricelist_id:
            return {'type': 'ir.actions.act_window_close'}
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'cable.pricelist',
            'res_id': self.pricelist_id.id,
            'view_mode': 'form',
        }
    
    def action_back(self):
        """Retour à l'étape précédente"""
        self.state = 'draft'
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
