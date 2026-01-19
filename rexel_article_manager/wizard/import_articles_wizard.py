# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import logging

_logger = logging.getLogger(__name__)

# Patch pour corriger l'erreur 'biltinId' dans openpyxl
# Cette erreur survient avec certains fichiers Excel créés avec des versions récentes d'Excel
# Le bug est une faute de frappe 'biltinId' au lieu de 'builtinId' dans les métadonnées du fichier
def _patch_openpyxl():
    try:
        from openpyxl.styles.named_styles import _NamedCellStyle
        _original_init = _NamedCellStyle.__init__
        
        def _patched_init(self, *args, **kwargs):
            # Supprimer 'biltinId' s'il est présent (c'est une faute de frappe pour 'builtinId')
            kwargs.pop('biltinId', None)
            return _original_init(self, *args, **kwargs)
        
        _NamedCellStyle.__init__ = _patched_init
        _logger.debug("Patch openpyxl biltinId appliqué avec succès")
    except Exception as e:
        _logger.warning(f"Impossible d'appliquer le patch openpyxl: {e}")

# Appliquer le patch au chargement du module
_patch_openpyxl()

import openpyxl
import requests


class ImportArticlesWizard(models.TransientModel):
    _name = 'import.articles.wizard'
    _description = 'Import Articles Rexel Cloud (Excel ou API)'

    # ========== CHOIX DU MODE ==========
    import_mode = fields.Selection([
        ('excel', 'Importer depuis un fichier Excel'),
        ('api', 'Importer depuis l\'API Rexel Cloud'),
    ], string='Mode d\'import', required=True, default='excel')
    
    # ========== IMPORT EXCEL ==========
    excel_file = fields.Binary(string='Fichier Excel', attachment=False)
    excel_filename = fields.Char(string='Nom du fichier')
    
    # ========== IMPORT API ==========
    use_api_config = fields.Boolean(
        string='Utiliser configuration',
        default=True,
        help='Utiliser les paramètres de la configuration Rexel'
    )
    api_customer_id = fields.Char(string='N° client Rexel')
    api_customer_scope = fields.Char(string='Mot client')
    
    # ========== OPTIONS D'IMPORT ==========
    import_mode_articles = fields.Selection([
        ('all', 'Créer nouveaux + Mettre à jour existants'),
        ('new_only', 'Créer uniquement les nouveaux articles'),
        ('update_only', 'Mettre à jour uniquement les existants'),
    ], string='Mode d\'import', default='all', required=True,
       help='Choisir quels articles traiter lors de l\'import')
    
    # Ancien champ conservé pour compatibilité (mais masqué)
    update_existing = fields.Boolean(
        string='Mettre à jour les articles existants',
        default=True,
        help='Met à jour les articles existants au lieu de les ignorer'
    )
    
    # ========== CHAMPS À METTRE À JOUR ==========
    update_prices = fields.Boolean(
        string='💰 Prix (base, net, vente, remise)',
        default=True,
        help='Mettre à jour les prix lors de l\'import'
    )
    
    update_designation = fields.Boolean(
        string='📝 Désignation',
        default=True,
        help='Mettre à jour la désignation'
    )
    
    update_famille = fields.Boolean(
        string='📁 Famille / Sous-famille / Fonction',
        default=True,
        help='Mettre à jour les informations de classification'
    )
    
    update_d3e = fields.Boolean(
        string='♻️ Informations D3E / DEEE',
        default=True,
        help='Mettre à jour les informations éco-contribution'
    )
    
    update_conditionnement = fields.Boolean(
        string='📦 Conditionnement / Unité',
        default=True,
        help='Mettre à jour le conditionnement et l\'unité de mesure'
    )
    
    update_dates = fields.Boolean(
        string='📅 Dates (tarif, péremption)',
        default=True,
        help='Mettre à jour les dates'
    )
    
    update_other = fields.Boolean(
        string='🔧 Autres informations',
        default=True,
        help='Mettre à jour les autres champs (EAN, ecoscore, URL image, etc.)'
    )
    
    create_products = fields.Boolean(
        string='Créer automatiquement les produits Odoo',
        default=False,
        help='Crée automatiquement les produits Odoo pour les nouveaux articles'
    )
    
    create_families = fields.Boolean(
        string='Créer automatiquement les familles',
        default=True,
        help='Créer automatiquement l\'arborescence Famille/Sous-famille/Fonction'
    )
    
    # ========== STATISTIQUES ==========
    articles_imported = fields.Integer(string='Articles importés', readonly=True)
    articles_updated = fields.Integer(string='Articles mis à jour', readonly=True)
    articles_skipped = fields.Integer(string='Articles ignorés', readonly=True)
    families_created = fields.Integer(string='Familles créées', readonly=True)
    import_log = fields.Text(string='Log d\'import', readonly=True)
    
    def _get_fields_to_update(self):
        """Retourne la liste des champs à mettre à jour selon les options"""
        fields_to_update = set()
        
        if self.update_prices:
            fields_to_update.update(['prix_base', 'prix_net', 'prix_vente', 'remise'])
        
        if self.update_designation:
            fields_to_update.add('designation')
        
        if self.update_famille:
            fields_to_update.update([
                'famille_code', 'famille_libelle',
                'sous_famille_code', 'sous_famille_libelle',
                'fonction_code', 'fonction_libelle',
                'family_node_id'
            ])
        
        if self.update_d3e:
            fields_to_update.update([
                'ref_d3e', 'libelle_d3e', 'code_d3e', 
                'unite_d3e', 'montant_d3e'
            ])
        
        if self.update_conditionnement:
            fields_to_update.update(['conditionnement', 'unite_mesure'])
        
        if self.update_dates:
            fields_to_update.update(['date_tarif', 'date_peremption'])
        
        if self.update_other:
            fields_to_update.update([
                'code_ean13', 'ecoscore', 'selection_durable', 
                'url_image', 'reference_rexel', 'code_lidic_and_ref'
            ])
        
        # Toujours mettre à jour ces champs d'identification
        fields_to_update.update([
            'trigramme_fabricant', 'fabricant_libelle', 'import_source'
        ])
        
        return fields_to_update

    @api.onchange('import_mode')
    def _onchange_import_mode(self):
        """Charge la configuration quand on passe en mode API"""
        if self.import_mode == 'api' and self.use_api_config:
            config = self.env['rexel.config'].get_config()
            self.api_customer_id = config.customer_id
            self.api_customer_scope = config.customer_scope

    def action_import(self):
        """Lance l'import selon le mode choisi"""
        self.ensure_one()
        
        if self.import_mode == 'excel':
            return self._import_from_excel()
        else:
            return self._import_from_api()

    def _import_from_excel(self):
        """Import depuis un fichier Excel"""
        if not self.excel_file:
            raise UserError(_('Veuillez sélectionner un fichier Excel.'))
        
        try:
            # Décoder le fichier Excel
            file_data = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(io.BytesIO(file_data))
            ws = wb.active
            
            # Lire les en-têtes (première ligne)
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # Mapping des colonnes Excel vers les champs Odoo
            field_mapping = self._get_excel_field_mapping()
            
            articles_imported = 0
            articles_updated = 0
            articles_skipped = 0
            families_created = 0
            log_messages = []
            skipped_reasons = {
                'no_reference': [],      # Pas de référence fabricant
                'existing_no_update': [], # Existe déjà et option MAJ désactivée
                'error': [],              # Erreur lors du traitement
            }
            
            # Cache pour les familles créées
            family_cache = {}
            
            # Parcourir les lignes de données
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Construire le dictionnaire de valeurs
                    values = {}
                    famille_info = {}
                    
                    for col_idx, cell_value in enumerate(row):
                        if col_idx < len(headers):
                            excel_col = headers[col_idx]
                            if excel_col in field_mapping:
                                odoo_field = field_mapping[excel_col]
                                if cell_value is not None:
                                    # Stocker les infos de famille séparément
                                    if odoo_field in ['famille_code', 'famille_libelle', 
                                                      'sous_famille_code', 'sous_famille_libelle',
                                                      'fonction_code', 'fonction_libelle']:
                                        famille_info[odoo_field] = str(cell_value).strip() if cell_value else ''
                                    values[odoo_field] = self._convert_value(odoo_field, cell_value)
                    
                    # Vérifier qu'on a au minimum la référence
                    if not values.get('reference_fabricant'):
                        articles_skipped += 1
                        # Essayer de trouver une info identifiable dans la ligne
                        row_info = f"ligne {row_idx}"
                        if values.get('designation'):
                            row_info += f" ({values['designation'][:30]}...)" if len(str(values.get('designation', ''))) > 30 else f" ({values.get('designation')})"
                        skipped_reasons['no_reference'].append(row_info)
                        continue
                    
                    # ========== CALCUL AUTOMATIQUE DE LA REMISE ==========
                    # Si la remise n'est pas fournie, la calculer à partir des prix
                    if not values.get('remise') and values.get('prix_base') and values.get('prix_net'):
                        prix_base = float(values['prix_base'])
                        prix_net = float(values['prix_net'])
                        if prix_base > 0:
                            remise_calculee = ((prix_base - prix_net) / prix_base) * 100
                            values['remise'] = round(remise_calculee, 2)
                    
                    # Créer l'arborescence des familles si demandé
                    if self.create_families and famille_info:
                        family_node_id, created_count = self._get_or_create_family_hierarchy(famille_info, family_cache)
                        if family_node_id:
                            values['family_node_id'] = family_node_id
                        families_created += created_count
                    
                    # Ajouter la source d'import
                    values['import_source'] = 'excel'
                    
                    # ========== RECHERCHE ARTICLE EXISTANT ==========
                    # Priorité 1: Référence Rexel (unique)
                    # Priorité 2: Référence fabricant + Trigramme fabricant
                    # Priorité 3: Code LIDIC AND REF (trigramme + référence concaténés)
                    existing = None
                    search_key = ""
                    
                    # Priorité 1: Référence Rexel (si disponible)
                    if values.get('reference_rexel'):
                        existing = self.env['rexel.article'].search([
                            ('reference_rexel', '=', values['reference_rexel'])
                        ], limit=1)
                        if existing:
                            search_key = f"REF_REXEL:{values['reference_rexel']}"
                    
                    # Priorité 2: Référence fabricant + Trigramme
                    if not existing and values.get('reference_fabricant') and values.get('trigramme_fabricant'):
                        existing = self.env['rexel.article'].search([
                            ('reference_fabricant', '=', values['reference_fabricant']),
                            ('trigramme_fabricant', '=', values['trigramme_fabricant'])
                        ], limit=1)
                        if existing:
                            search_key = f"{values['trigramme_fabricant']}:{values['reference_fabricant']}"
                    
                    # Priorité 3: Code LIDIC AND REF
                    if not existing and values.get('code_lidic_and_ref'):
                        existing = self.env['rexel.article'].search([
                            ('code_lidic_and_ref', '=', values['code_lidic_and_ref'])
                        ], limit=1)
                        if existing:
                            search_key = f"CODE:{values['code_lidic_and_ref']}"
                    
                    # Priorité 4 (fallback): Référence fabricant seule SI pas de trigramme fourni
                    if not existing and values.get('reference_fabricant') and not values.get('trigramme_fabricant'):
                        existing = self.env['rexel.article'].search([
                            ('reference_fabricant', '=', values['reference_fabricant'])
                        ], limit=1)
                        if existing:
                            search_key = f"REF:{values['reference_fabricant']} (sans trigramme)"
                    
                    # ========== TRAITEMENT SELON LE MODE D'IMPORT ==========
                    if existing:
                        # Article existant trouvé
                        if self.import_mode_articles == 'new_only':
                            # Mode "nouveaux uniquement" → ignorer les existants
                            articles_skipped += 1
                            skipped_reasons['existing_no_update'].append(
                                f"{values['reference_fabricant']} (mode nouveaux uniquement)"
                            )
                        elif self.import_mode_articles in ('all', 'update_only'):
                            # Mode "tous" ou "mise à jour uniquement" → mettre à jour
                            # Filtrer les champs à mettre à jour selon les options
                            fields_to_update = self._get_fields_to_update()
                            filtered_values = {k: v for k, v in values.items() if k in fields_to_update}
                            
                            # Enregistrer l'ancien prix dans l'historique si prix modifié
                            if self.update_prices and 'prix_net' in filtered_values:
                                if existing.prix_net != filtered_values.get('prix_net', 0):
                                    self.env['rexel.price.history'].create({
                                        'article_id': existing.id,
                                        'prix_base': existing.prix_base or existing.prix_net,
                                        'prix_net': filtered_values['prix_net'],
                                        'source': 'import',
                                        'date': fields.Datetime.now(),
                                    })
                            
                            if filtered_values:
                                existing.write(filtered_values)
                                articles_updated += 1
                            else:
                                articles_skipped += 1
                                skipped_reasons['existing_no_update'].append(
                                    f"{values['reference_fabricant']} (aucun champ à MAJ)"
                                )
                    else:
                        # Nouvel article
                        if self.import_mode_articles == 'update_only':
                            # Mode "mise à jour uniquement" → ignorer les nouveaux
                            articles_skipped += 1
                            skipped_reasons['new_skipped'] = skipped_reasons.get('new_skipped', [])
                            skipped_reasons['new_skipped'].append(
                                f"{values['reference_fabricant']} (mode MAJ uniquement)"
                            )
                        elif self.import_mode_articles in ('all', 'new_only'):
                            # Mode "tous" ou "nouveaux uniquement" → créer
                            article = self.env['rexel.article'].create(values)
                            articles_imported += 1
                            
                            # Créer le produit Odoo si demandé
                            if self.create_products:
                                try:
                                    article.action_create_product()
                                except Exception as e:
                                    log_messages.append(f"⚠️ Erreur création produit {article.reference_fabricant}: {str(e)}")
                    
                except Exception as e:
                    articles_skipped += 1
                    ref_info = values.get('reference_fabricant', f'ligne {row_idx}')
                    skipped_reasons['error'].append(f"{ref_info}: {str(e)}")
                    _logger.error(f"Erreur import ligne {row_idx}: {str(e)}")
            
            # Mettre à jour les statistiques
            self.articles_imported = articles_imported
            self.articles_updated = articles_updated
            self.articles_skipped = articles_skipped
            self.families_created = families_created
            
            # ========== CONSTRUCTION DU LOG DÉTAILLÉ ==========
            log_lines = []
            log_lines.append("=" * 50)
            log_lines.append("📊 RÉSUMÉ DE L'IMPORT")
            log_lines.append("=" * 50)
            
            # Afficher le mode d'import
            mode_labels = {
                'all': 'Créer + Mettre à jour',
                'new_only': 'Nouveaux uniquement',
                'update_only': 'Mise à jour uniquement'
            }
            log_lines.append(f"📋 Mode: {mode_labels.get(self.import_mode_articles, self.import_mode_articles)}")
            
            # Afficher les champs mis à jour
            if self.import_mode_articles != 'new_only':
                updated_fields = []
                if self.update_prices: updated_fields.append("Prix")
                if self.update_designation: updated_fields.append("Désignation")
                if self.update_famille: updated_fields.append("Famille")
                if self.update_d3e: updated_fields.append("D3E")
                if self.update_conditionnement: updated_fields.append("Conditionnement")
                if self.update_dates: updated_fields.append("Dates")
                if self.update_other: updated_fields.append("Autres")
                if updated_fields:
                    log_lines.append(f"🔧 Champs MAJ: {', '.join(updated_fields)}")
            
            log_lines.append("")
            log_lines.append(f"✅ Articles créés: {articles_imported}")
            log_lines.append(f"🔄 Articles mis à jour: {articles_updated}")
            log_lines.append(f"⏭️ Articles ignorés: {articles_skipped}")
            if families_created > 0:
                log_lines.append(f"📁 Familles créées: {families_created}")
            
            # Détail des articles ignorés
            if articles_skipped > 0:
                log_lines.append("")
                log_lines.append("=" * 50)
                log_lines.append("⚠️ DÉTAIL DES ARTICLES IGNORÉS")
                log_lines.append("=" * 50)
                
                if skipped_reasons['no_reference']:
                    log_lines.append("")
                    log_lines.append(f"❌ Sans référence fabricant ({len(skipped_reasons['no_reference'])}):")
                    for item in skipped_reasons['no_reference'][:20]:  # Limiter à 20
                        log_lines.append(f"   • {item}")
                    if len(skipped_reasons['no_reference']) > 20:
                        log_lines.append(f"   ... et {len(skipped_reasons['no_reference']) - 20} autres")
                
                if skipped_reasons['existing_no_update']:
                    log_lines.append("")
                    log_lines.append(f"🔒 Existants non mis à jour ({len(skipped_reasons['existing_no_update'])}):")
                    for ref in skipped_reasons['existing_no_update'][:20]:  # Limiter à 20
                        log_lines.append(f"   • {ref}")
                    if len(skipped_reasons['existing_no_update']) > 20:
                        log_lines.append(f"   ... et {len(skipped_reasons['existing_no_update']) - 20} autres")
                
                if skipped_reasons.get('new_skipped'):
                    log_lines.append("")
                    log_lines.append(f"🆕 Nouveaux non créés ({len(skipped_reasons['new_skipped'])}):")
                    for ref in skipped_reasons['new_skipped'][:20]:  # Limiter à 20
                        log_lines.append(f"   • {ref}")
                    if len(skipped_reasons['new_skipped']) > 20:
                        log_lines.append(f"   ... et {len(skipped_reasons['new_skipped']) - 20} autres")
                
                if skipped_reasons['error']:
                    log_lines.append("")
                    log_lines.append(f"🚫 Erreurs de traitement ({len(skipped_reasons['error'])}):")
                    for err in skipped_reasons['error'][:20]:  # Limiter à 20
                        log_lines.append(f"   • {err}")
                    if len(skipped_reasons['error']) > 20:
                        log_lines.append(f"   ... et {len(skipped_reasons['error']) - 20} autres")
            
            # Ajouter les messages d'erreur supplémentaires
            if log_messages:
                log_lines.append("")
                log_lines.append("=" * 50)
                log_lines.append("📝 AUTRES MESSAGES")
                log_lines.append("=" * 50)
                log_lines.extend(log_messages)
            
            self.import_log = '\n'.join(log_lines)
            
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'import.articles.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
                'context': {'show_results': True},
            }
            
        except Exception as e:
            raise UserError(_(f'Erreur lors de l\'import Excel: {str(e)}'))

    def _get_or_create_family_hierarchy(self, famille_info, cache):
        """
        Créer ou récupérer l'arborescence Famille > Sous-famille > Fonction
        Retourne (famille_id, nombre_créées)
        """
        Family = self.env['rexel.product.family']
        created_count = 0
        
        famille_code = famille_info.get('famille_code', '').strip()
        famille_libelle = famille_info.get('famille_libelle', '').strip()
        sous_famille_code = famille_info.get('sous_famille_code', '').strip()
        sous_famille_libelle = famille_info.get('sous_famille_libelle', '').strip()
        fonction_code = famille_info.get('fonction_code', '').strip()
        fonction_libelle = famille_info.get('fonction_libelle', '').strip()
        
        # Niveau 1: Famille
        famille_id = None
        if famille_code or famille_libelle:
            cache_key = f"famille_{famille_code}_{famille_libelle}"
            if cache_key in cache:
                famille_id = cache[cache_key]
            else:
                # Chercher par code d'abord, puis par nom
                famille = None
                if famille_code:
                    famille = Family.search([('code', '=', famille_code), ('level', '=', 'famille')], limit=1)
                if not famille and famille_libelle:
                    famille = Family.search([('name', '=', famille_libelle), ('level', '=', 'famille')], limit=1)
                
                if not famille:
                    famille = Family.create({
                        'name': famille_libelle or famille_code,
                        'code': famille_code or False,
                        'level': 'famille',
                    })
                    created_count += 1
                    _logger.info(f"Famille créée: {famille.name}")
                
                famille_id = famille.id
                cache[cache_key] = famille_id
        
        # Niveau 2: Sous-famille
        sous_famille_id = None
        if (sous_famille_code or sous_famille_libelle) and famille_id:
            cache_key = f"sousfamille_{famille_id}_{sous_famille_code}_{sous_famille_libelle}"
            if cache_key in cache:
                sous_famille_id = cache[cache_key]
            else:
                sous_famille = None
                if sous_famille_code:
                    sous_famille = Family.search([
                        ('code', '=', sous_famille_code), 
                        ('level', '=', 'sous_famille'),
                        ('parent_id', '=', famille_id)
                    ], limit=1)
                if not sous_famille and sous_famille_libelle:
                    sous_famille = Family.search([
                        ('name', '=', sous_famille_libelle), 
                        ('level', '=', 'sous_famille'),
                        ('parent_id', '=', famille_id)
                    ], limit=1)
                
                if not sous_famille:
                    sous_famille = Family.create({
                        'name': sous_famille_libelle or sous_famille_code,
                        'code': sous_famille_code or False,
                        'level': 'sous_famille',
                        'parent_id': famille_id,
                    })
                    created_count += 1
                    _logger.info(f"Sous-famille créée: {sous_famille.name}")
                
                sous_famille_id = sous_famille.id
                cache[cache_key] = sous_famille_id
        
        # Niveau 3: Fonction
        fonction_id = None
        parent_for_fonction = sous_famille_id or famille_id
        if (fonction_code or fonction_libelle) and parent_for_fonction:
            cache_key = f"fonction_{parent_for_fonction}_{fonction_code}_{fonction_libelle}"
            if cache_key in cache:
                fonction_id = cache[cache_key]
            else:
                fonction = None
                if fonction_code:
                    fonction = Family.search([
                        ('code', '=', fonction_code), 
                        ('level', '=', 'fonction'),
                        ('parent_id', '=', parent_for_fonction)
                    ], limit=1)
                if not fonction and fonction_libelle:
                    fonction = Family.search([
                        ('name', '=', fonction_libelle), 
                        ('level', '=', 'fonction'),
                        ('parent_id', '=', parent_for_fonction)
                    ], limit=1)
                
                if not fonction:
                    fonction = Family.create({
                        'name': fonction_libelle or fonction_code,
                        'code': fonction_code or False,
                        'level': 'fonction',
                        'parent_id': parent_for_fonction,
                    })
                    created_count += 1
                    _logger.info(f"Fonction créée: {fonction.name}")
                
                fonction_id = fonction.id
                cache[cache_key] = fonction_id
        
        # Retourner le niveau le plus bas disponible
        final_id = fonction_id or sous_famille_id or famille_id
        return final_id, created_count

    def _import_from_api(self):
        """Import depuis l'API Rexel Cloud"""
        if not self.api_customer_id:
            raise UserError(_('Le numéro client Rexel est requis pour l\'import API.'))
        
        config = self.env['rexel.config'].get_config()
        
        if not config.api_enabled:
            raise UserError(_('L\'API n\'est pas activée dans la configuration.'))
        
        try:
            # TODO: Appeler l'API ProductPrice pour récupérer tous les articles
            # Note: L'API Rexel nécessite de connaître les références à l'avance
            # Il faudra probablement combiner avec un export Excel initial
            
            raise UserError(_(
                'L\'import complet via API n\'est pas encore implémenté.\n\n'
                'L\'API Rexel nécessite de connaître les références produits à l\'avance.\n'
                'Utilisez d\'abord un import Excel, puis la fonction de mise à jour des prix via API.\n\n'
                'Ou utilisez la fonction "Import par Référence" pour importer des références individuelles.'
            ))
            
        except Exception as e:
            raise UserError(_(f'Erreur lors de l\'import API: {str(e)}'))

    def _get_excel_field_mapping(self):
        """
        Mapping des colonnes Excel cloud Rexel vers les champs Odoo
        Supporte plusieurs formats d'export Rexel
        """
        return {
            # ========== FORMAT EXPORT REXEL CLOUD (format principal) ==========
            # Ces en-têtes correspondent au format d'export standard de Rexel
            'Référence': 'reference_fabricant',  # Alias pour compatibilité
            'Référence Rexel': 'reference_rexel',
            'Code Lidic + Réf.': 'code_lidic_and_ref',
            'Désignation': 'designation',
            'Code EAN13': 'code_ean13',
            'Code Lidic': 'trigramme_fabricant',
            'Libellé Lidic': 'fabricant_libelle',
            'Prix de Base': 'prix_base',
            'Prix Net': 'prix_net',
            'Prix de Vente': 'prix_vente',
            'Date Tarif': 'date_tarif',
            'Date péremption': 'date_peremption',
            'Conditionnement': 'conditionnement',
            'Famille': 'famille_code',
            'Sous Famille': 'sous_famille_code',
            'Fonction': 'fonction_code',
            'Libellé Famille': 'famille_libelle',
            'Libellé Sous Famille': 'sous_famille_libelle',
            'Libellé Fonction': 'fonction_libelle',
            'Réf. DEEE': 'ref_d3e',
            'Libellé DEEE': 'libelle_d3e',
            'Code DEEE': 'code_d3e',
            'Unité DEEE': 'unite_d3e',
            'Montant DEEE': 'montant_d3e',
            
            # ========== FORMAT RÉEXPORT ODOO (format préféré) ==========
            'Référence Fabricant': 'reference_fabricant',
            'Trigramme Fabricant': 'trigramme_fabricant',
            'Prix Base': 'prix_base',
            'Prix Net': 'prix_net',
            'Remise %': 'remise',
            'Unité Mesure': 'unite_mesure',
            'Code EAN': 'code_ean13',
            'Montant D3E': 'montant_d3e',
            'Unité D3E': 'unite_d3e',
            'Sous-Famille': 'sous_famille_libelle',
            'Code Famille': 'famille_code',
            'Code Sous-Famille': 'sous_famille_code',
            'Code Fonction': 'fonction_code',
            
            # ========== FORMAT INTERNE (colonnes techniques) ==========
            'REF': 'reference_fabricant',
            'REF_REXEL': 'reference_rexel',
            'CODE_LIDIC_AND_REF': 'code_lidic_and_ref',
            'DESIGNATION': 'designation',
            'CODE_EAN13': 'code_ean13',
            'CODE_LIDIC': 'trigramme_fabricant',
            'LIBELLE_FABRICANT': 'fabricant_libelle',
            'PRIX_BASE': 'prix_base',
            'PRIX_NET': 'prix_net',
            'PRIX_VENTE': 'prix_vente',
            'REMISE': 'remise',
            'DATE_TARIF': 'date_tarif',
            'DATE_PEREMPTION': 'date_peremption',
            'CONDITIONNEMENT': 'conditionnement',
            'UOM': 'unite_mesure',
            'FAMILLE': 'famille_code',
            'SOUS_FAMILLE': 'sous_famille_code',
            'FONCTION': 'fonction_code',
            'LIBELLE_FAMILLE': 'famille_libelle',
            'LIBELLE_SOUS_FAMILLE': 'sous_famille_libelle',
            'LIBELLE_FONCTION': 'fonction_libelle',
            'REF_D3E': 'ref_d3e',
            'LIBELLE_D3E': 'libelle_d3e',
            'CODE_D3E': 'code_d3e',
            'UNITE_D3E': 'unite_d3e',
            'MONTANT_D3E': 'montant_d3e',
            'ECOSCORE': 'ecoscore',
            'SELECTION_DURABLE': 'selection_durable',
            'URL_IMAGE': 'url_image',
            
            # ========== FORMAT TECHNIQUE (noms de champs Odoo) ==========
            'reference_fabricant': 'reference_fabricant',
            'reference_rexel': 'reference_rexel',
            'trigramme_fabricant': 'trigramme_fabricant',
            'designation': 'designation',
            'code_ean13': 'code_ean13',
            'fabricant_libelle': 'fabricant_libelle',
            'prix_base': 'prix_base',
            'prix_net': 'prix_net',
            'prix_vente': 'prix_vente',
            'remise': 'remise',
            'date_tarif': 'date_tarif',
            'date_peremption': 'date_peremption',
            'conditionnement': 'conditionnement',
            'unite_mesure': 'unite_mesure',
            'famille_code': 'famille_code',
            'sous_famille_code': 'sous_famille_code',
            'fonction_code': 'fonction_code',
            'famille_libelle': 'famille_libelle',
            'sous_famille_libelle': 'sous_famille_libelle',
            'fonction_libelle': 'fonction_libelle',
            'montant_d3e': 'montant_d3e',
            'unite_d3e': 'unite_d3e',
        }

    def _convert_value(self, field_name, value):
        """Convertit une valeur selon le type de champ"""
        if value is None or value == '':
            return False
        
        # Champs booléens
        if field_name == 'selection_durable':
            if isinstance(value, bool):
                return value
            if isinstance(value, str):
                return value.lower() in ['true', '1', 'oui', 'yes', 'o', 'y']
            return bool(value)
        
        # Champs numériques
        if field_name in ['prix_base', 'prix_net', 'prix_vente', 'remise', 'unite_d3e', 'montant_d3e']:
            try:
                return float(value)
            except (ValueError, TypeError):
                return 0.0
        
        # Champs date
        if field_name in ['date_tarif', 'date_peremption']:
            if value is None or value == '':
                return False
            # Si c'est déjà un objet date/datetime
            if hasattr(value, 'strftime'):
                return value.strftime('%Y-%m-%d')
            # Si c'est une chaîne
            if isinstance(value, str):
                value = value.strip()
                if not value:
                    return False
                # Essayer différents formats de date
                from datetime import datetime
                date_formats = [
                    '%d/%m/%Y',      # 15/12/2025
                    '%Y-%m-%d',      # 2025-12-15
                    '%d-%m-%Y',      # 15-12-2025
                    '%d.%m.%Y',      # 15.12.2025
                    '%Y/%m/%d',      # 2025/12/15
                ]
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(value, fmt)
                        return parsed_date.strftime('%Y-%m-%d')
                    except ValueError:
                        continue
                # Si aucun format ne correspond, retourner False
                _logger.warning(f"Format de date non reconnu: {value}")
                return False
            return False
        
        # Champs texte
        return str(value).strip()

    def action_update_prices_from_api(self):
        """Met à jour les prix des articles existants depuis l'API"""
        config = self.env['rexel.config'].get_config()
        
        if not config.api_enabled:
            raise UserError(_('L\'API n\'est pas activée.'))
        
        # Récupérer tous les articles
        articles = self.env['rexel.article'].search([])
        
        if not articles:
            raise UserError(_('Aucun article à mettre à jour.'))
        
        # Appeler l'API par batch de 50 articles
        batch_size = 50
        updated_count = 0
        
        for i in range(0, len(articles), batch_size):
            batch = articles[i:i+batch_size]
            prices = config.get_product_prices_from_api(batch)
            
            for article_id, price_data in prices.items():
                article = self.env['rexel.article'].browse(article_id)
                
                # Enregistrer dans l'historique si le prix change
                if article.prix_net != price_data['prix_net']:
                    self.env['rexel.price.history'].create({
                        'article_id': article.id,
                        'prix_base': article.prix_base or article.prix_net,
                        'prix_net': price_data['prix_net'],
                        'source': 'rexel_api',
                        'date': fields.Datetime.now(),
                    })
                
                article.write(price_data)
                updated_count += 1
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Mise à jour terminée'),
                'message': _(f'{updated_count} articles mis à jour depuis l\'API Rexel.'),
                'type': 'success',
                'sticky': False,
            }
        }
