# -*- coding: utf-8 -*-

import base64
import io
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import logging

_logger = logging.getLogger(__name__)


class ExportNaviwestWizard(models.TransientModel):
    _name = 'export.naviwest.wizard'
    _description = 'Assistant d\'export vers Naviwest'

    export_type = fields.Selection([
        ('beg', 'BEG'),
        ('niedax', 'NIEDAX'),
        ('cables', 'CÂBLES'),
    ], string='Type d\'export', required=True, default='beg')
    
    # Filtres avancés
    filter_field = fields.Selection([
        ('famille', 'Famille'),
        ('sous_famille', 'Sous-famille'),
        ('fonction', 'Fonction'),
        ('lidic_libelle', 'Fabricant (Lidic)'),
        ('gamme_vente_libelle', 'Gamme de vente'),
    ], string='Filtrer par')
    
    filter_value = fields.Char(string='Valeur du filtre', help='Ex: CABLE, LEGRAND, etc.')
    
    # Ancien filtre (conservé pour compatibilité)
    filter_by_family = fields.Boolean(string='Filtrer par famille (ancien)', default=False)
    family_filter = fields.Char(string='Filtre famille (ancien)')
    
    article_ids = fields.Many2many('rexel.article', string='Articles à exporter')
    
    export_file = fields.Binary(string='Fichier exporté', readonly=True)
    export_filename = fields.Char(string='Nom du fichier', readonly=True)

    @api.model
    def default_get(self, fields_list):
        """Récupère les articles sélectionnés dans la vue liste"""
        res = super().default_get(fields_list)
        
        if self.env.context.get('active_ids'):
            res['article_ids'] = [(6, 0, self.env.context.get('active_ids'))]
        
        return res

    def action_export(self):
        """Lance l'export vers Naviwest en utilisant le template Excel"""
        self.ensure_one()
        
        if not self.article_ids:
            raise UserError(_('Veuillez sélectionner au moins un article à exporter.'))
        
        # Récupérer la configuration
        config = self.env['rexel.config'].get_config()
        
        # Chemins des fichiers template depuis la configuration
        template_files = {
            'beg': config.template_beg,
            'niedax': config.template_niedax,
            'cables': config.template_cables,
        }
        
        template_path = template_files.get(self.export_type)
        
        if not template_path or not os.path.exists(template_path):
            # Message d'erreur détaillé
            type_names = {'beg': 'BEG', 'niedax': 'NIEDAX', 'cables': 'CÂBLES'}
            type_name = type_names.get(self.export_type)
            
            error_msg = f"""Fichier template {type_name} non trouvé.

Chemin configuré : {template_path or '(non configuré)'}
Fichier existe : {os.path.exists(template_path) if template_path else 'N/A'}

Pour corriger :
1. Menu → Configuration → Configuration Rexel
2. Vérifier le chemin du template {type_name}
3. Cliquer sur "Vérifier les templates"
4. S'assurer que le fichier existe bien à l'emplacement indiqué

Chemin attendu : C:\\Users\\dimitri\\Documents\\fichier odoo\\Export naviwest et quickdevis\\MAJ Base_Article - {type_name} avec formules V1.xlsx"""
            raise UserError(_(error_msg))
        
        try:
            # Charger le fichier template
            wb = load_workbook(template_path)
            
            # La feuille principale est 'Article'
            if 'Article' in wb.sheetnames:
                ws = wb['Article']
            else:
                ws = wb.active
            
            # Supprimer les données existantes (garder la ligne d'en-tête)
            # On part de la ligne 2
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            
            # Remplir avec les nouveaux articles
            current_row = 2
            
            # Appliquer les filtres
            articles = self.article_ids
            if self.filter_field and self.filter_value:
                # Nouveau système de filtrage
                articles = articles.filtered(
                    lambda a: a[self.filter_field] and 
                    self.filter_value.upper() in str(a[self.filter_field]).upper()
                )
            elif self.filter_by_family and self.family_filter:
                # Ancien système (compatibilité)
                articles = articles.filtered(
                    lambda a: a.famille_libelle and 
                    self.family_filter.upper() in a.famille_libelle.upper()
                )
            
            if not articles:
                raise UserError(_('Aucun article correspondant aux filtres sélectionnés.'))
            
            for article in articles:
                # Col 1: N° (Référence)
                ws.cell(current_row, 1, article.reference_fabricant or '')
                
                # Col 2: Description
                ws.cell(current_row, 2, article.designation or '')
                
                # Col 3: Description de recherche (50 premiers caractères)
                desc_recherche = article.designation[:50] if article.designation else ''
                ws.cell(current_row, 3, desc_recherche)
                
                # Col 4: Unité de base
                ws.cell(current_row, 4, article.unite_mesure or 'U')
                
                # Col 5: Groupe compta. stock (vide ou valeur du template)
                ws.cell(current_row, 5, 'MP')  # Valeur par défaut du template
                
                # Col 6: N° emplacement (vide)
                ws.cell(current_row, 6, '')
                
                # Col 7: Code raccourci axe 3 (Code Lidic)
                ws.cell(current_row, 7, article.trigramme_fabricant or '')
                
                # Col 8: Coût unitaire (FORMULE - on la copie depuis la ligne 2 du template d'origine)
                # La formule fait référence à d'autres feuilles, on la garde
                # On met le prix net directement pour l'instant
                ws.cell(current_row, 8, article.prix_net or 0.0)
                
                # Col 9: Coût étude (FORMULE =H2)
                # On peut mettre une formule qui référence la colonne H
                ws.cell(current_row, 9).value = f'=H{current_row}'
                
                # Col 10: Dernier coût direct
                ws.cell(current_row, 10, article.prix_net or 0.0)
                
                # Col 11: N° fournisseur
                ws.cell(current_row, 11, 'REXEL')
                
                # Col 12: Référence fournisseur
                ws.cell(current_row, 12, article.reference_rexel or article.reference_fabricant or '')
                
                # Col 13: Poids brut (vide)
                ws.cell(current_row, 13, '')
                
                # Col 14: Poids net (vide)
                ws.cell(current_row, 14, '')
                
                # Col 15: Groupe compta. produit (vide)
                ws.cell(current_row, 15, '')
                
                # Col 16: Unité de vente
                ws.cell(current_row, 16, article.unite_mesure or 'U')
                
                # Col 17: Unité d'achat
                ws.cell(current_row, 17, article.unite_mesure or 'U')
                
                # Col 18: Procédure d'appro. (vide)
                ws.cell(current_row, 18, '')
                
                # Col 19: Code Famille
                code_famille = self._get_code_famille()
                ws.cell(current_row, 19, code_famille)
                
                # Col 20: Stocks
                ws.cell(current_row, 20, 0)
                
                current_row += 1
            
            # Sauvegarder le fichier
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Préparer le nom du fichier
            export_type_name = dict(self._fields['export_type'].selection).get(self.export_type)
            filename = f"MAJ_Base_Article_{export_type_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            # Encoder en base64
            excel_data = output.getvalue()
            export_file = base64.b64encode(excel_data)
            
            # Sauvegarder le fichier dans le wizard
            self.write({
                'export_file': export_file,
                'export_filename': filename,
            })
            
            # Retourner directement l'action de téléchargement
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/export.naviwest.wizard/{self.id}/export_file/{filename}?download=true',
                'target': 'self',
            }
            
        except Exception as e:
            _logger.error(f"Erreur lors de l'export Naviwest: {str(e)}")
            raise UserError(_('Erreur lors de l\'export : %s') % str(e))

    def _get_code_famille(self):
        """Détermine le code famille pour Naviwest selon le type d'export"""
        if self.export_type == 'beg':
            return '8'
        elif self.export_type == 'niedax':
            return '9'
        elif self.export_type == 'cables':
            return '1'
        return ''

    def action_download(self):
        """Télécharge le fichier exporté"""
        self.ensure_one()
        
        if not self.export_file:
            raise UserError(_('Aucun fichier à télécharger. Veuillez d\'abord lancer l\'export.'))
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/export.naviwest.wizard/{self.id}/export_file/{self.export_filename}?download=true',
            'target': 'self',
        }
