# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import logging

_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    _logger.warning("openpyxl n'est pas installé. L'export Excel ne sera pas disponible.")


class ExportRexelWizard(models.TransientModel):
    _name = 'export.rexel.wizard'
    _description = 'Assistant d\'export format Rexel'

    # Filtres de sélection
    famille_ids = fields.Many2many(
        'rexel.product.family',
        'export_rexel_wizard_famille_rel',
        'wizard_id', 'famille_id',
        string='Familles',
        domain=[('parent_id', '=', False)]
    )
    article_ids = fields.Many2many(
        'rexel.article',
        'export_rexel_wizard_article_rel',
        'wizard_id', 'article_id',
        string='Articles à exporter'
    )
    
    # Options
    export_all = fields.Boolean(
        string='Exporter tous les articles',
        default=False
    )
    include_obsolete = fields.Boolean(
        string='Inclure les articles obsolètes',
        default=False
    )
    
    # ===== COLONNES À EXPORTER =====
    col_reference_fabricant = fields.Boolean(string='Référence Fabricant', default=True)
    col_trigramme_fabricant = fields.Boolean(string='Trigramme Fabricant', default=True)
    col_fabricant_libelle = fields.Boolean(string='Fabricant (Libellé)', default=True)
    col_reference_rexel = fields.Boolean(string='Référence Rexel', default=True)
    col_designation = fields.Boolean(string='Désignation', default=True)
    col_prix_base = fields.Boolean(string='Prix Base', default=True)
    col_prix_net = fields.Boolean(string='Prix Net', default=True)
    col_remise = fields.Boolean(string='Remise %', default=True)
    col_unite_mesure = fields.Boolean(string='Unité Mesure', default=True)
    col_conditionnement = fields.Boolean(string='Conditionnement', default=True)
    col_code_ean = fields.Boolean(string='Code EAN', default=True)
    col_montant_d3e = fields.Boolean(string='Montant D3E', default=False)
    col_unite_d3e = fields.Boolean(string='Unité D3E', default=False)
    col_famille = fields.Boolean(string='Famille', default=True)
    col_sous_famille = fields.Boolean(string='Sous-Famille', default=True)
    col_fonction = fields.Boolean(string='Fonction', default=True)
    col_code_famille = fields.Boolean(string='Code Famille', default=False)
    col_code_sous_famille = fields.Boolean(string='Code Sous-Famille', default=False)
    col_code_fonction = fields.Boolean(string='Code Fonction', default=False)
    col_date_maj = fields.Boolean(string='Date MAJ Prix', default=False)
    col_obsolete = fields.Boolean(string='Obsolète', default=False)
    col_source = fields.Boolean(string='Source', default=False)
    
    # Résultat
    file_data = fields.Binary(string='Fichier', readonly=True)
    file_name = fields.Char(string='Nom du fichier', readonly=True)
    state = fields.Selection([
        ('draft', 'Configuration'),
        ('done', 'Terminé')
    ], default='draft')
    
    # Statistiques
    articles_exported = fields.Integer(string='Articles exportés', readonly=True)

    def action_select_all_columns(self):
        """Sélectionner toutes les colonnes"""
        self.write({
            'col_reference_fabricant': True,
            'col_trigramme_fabricant': True,
            'col_fabricant_libelle': True,
            'col_reference_rexel': True,
            'col_designation': True,
            'col_prix_base': True,
            'col_prix_net': True,
            'col_remise': True,
            'col_unite_mesure': True,
            'col_conditionnement': True,
            'col_code_ean': True,
            'col_montant_d3e': True,
            'col_unite_d3e': True,
            'col_famille': True,
            'col_sous_famille': True,
            'col_fonction': True,
            'col_code_famille': True,
            'col_code_sous_famille': True,
            'col_code_fonction': True,
            'col_date_maj': True,
            'col_obsolete': True,
            'col_source': True,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.rexel.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
    
    def action_select_minimal_columns(self):
        """Sélectionner les colonnes minimales"""
        self.write({
            'col_reference_fabricant': True,
            'col_trigramme_fabricant': True,
            'col_fabricant_libelle': True,
            'col_reference_rexel': True,
            'col_designation': True,
            'col_prix_base': True,
            'col_prix_net': True,
            'col_remise': True,
            'col_unite_mesure': True,
            'col_conditionnement': False,
            'col_code_ean': False,
            'col_montant_d3e': False,
            'col_unite_d3e': False,
            'col_famille': True,
            'col_sous_famille': False,
            'col_fonction': False,
            'col_code_famille': False,
            'col_code_sous_famille': False,
            'col_code_fonction': False,
            'col_date_maj': False,
            'col_obsolete': False,
            'col_source': False,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.rexel.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    @api.onchange('famille_ids')
    def _onchange_famille_ids(self):
        """Filtrer les articles selon les familles sélectionnées"""
        if self.famille_ids:
            # Récupérer toutes les familles (y compris sous-familles et fonctions)
            all_family_ids = self._get_all_subfamilies(self.famille_ids.ids)
            domain = [('family_node_id', 'in', all_family_ids)]
            if not self.include_obsolete:
                domain.append(('is_obsolete', '=', False))
            return {'domain': {'article_ids': domain}}
        return {'domain': {'article_ids': []}}

    def _get_all_subfamilies(self, family_ids):
        """Récupère récursivement toutes les sous-familles"""
        Family = self.env['rexel.product.family']
        all_ids = list(family_ids)
        children = Family.search([('parent_id', 'in', family_ids)])
        if children:
            all_ids.extend(self._get_all_subfamilies(children.ids))
        return all_ids

    def action_export(self):
        """Exporter les articles au format Rexel Excel"""
        self.ensure_one()
        
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("La bibliothèque openpyxl n'est pas installée. "
                            "Veuillez l'installer avec: pip install openpyxl"))
        
        # Déterminer les articles à exporter
        if self.export_all:
            domain = []
            if not self.include_obsolete:
                domain.append(('is_obsolete', '=', False))
            articles = self.env['rexel.article'].search(domain)
        elif self.article_ids:
            articles = self.article_ids
        elif self.famille_ids:
            all_family_ids = self._get_all_subfamilies(self.famille_ids.ids)
            domain = [('family_node_id', 'in', all_family_ids)]
            if not self.include_obsolete:
                domain.append(('is_obsolete', '=', False))
            articles = self.env['rexel.article'].search(domain)
        else:
            raise UserError(_("Veuillez sélectionner des articles ou des familles à exporter, "
                            "ou cochez 'Exporter tous les articles'."))
        
        if not articles:
            raise UserError(_("Aucun article à exporter."))
        
        # Créer le fichier Excel
        excel_data = self._create_rexel_excel(articles)
        
        # Mettre à jour l'assistant
        self.write({
            'file_data': base64.b64encode(excel_data),
            'file_name': f'export_rexel_{fields.Date.today()}.xlsx',
            'state': 'done',
            'articles_exported': len(articles)
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.rexel.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _get_selected_columns(self):
        """Retourne la liste des colonnes sélectionnées avec leurs propriétés"""
        all_columns = [
            ('col_reference_fabricant', 'Référence Fabricant', 'reference_fabricant', 20),
            ('col_trigramme_fabricant', 'Trigramme Fabricant', 'trigramme_fabricant', 15),
            ('col_fabricant_libelle', 'Fabricant', 'fabricant_libelle', 25),
            ('col_reference_rexel', 'Référence Rexel', 'reference_rexel', 20),
            ('col_designation', 'Désignation', 'designation', 50),
            ('col_prix_base', 'Prix Base', 'prix_base', 12),
            ('col_prix_net', 'Prix Net', 'prix_net', 12),
            ('col_remise', 'Remise %', 'remise', 10),
            ('col_unite_mesure', 'Unité Mesure', 'unite_mesure', 12),
            ('col_conditionnement', 'Conditionnement', 'conditionnement', 15),
            ('col_code_ean', 'Code EAN', 'code_ean13', 15),
            ('col_montant_d3e', 'Montant D3E', 'montant_d3e', 12),
            ('col_unite_d3e', 'Unité D3E', 'unite_d3e', 10),
            ('col_famille', 'Famille', 'famille_libelle', 30),
            ('col_sous_famille', 'Sous-Famille', 'sous_famille_libelle', 30),
            ('col_fonction', 'Fonction', 'fonction_libelle', 30),
            ('col_code_famille', 'Code Famille', 'famille_code', 15),
            ('col_code_sous_famille', 'Code Sous-Famille', 'sous_famille_code', 15),
            ('col_code_fonction', 'Code Fonction', 'fonction_code', 15),
            ('col_date_maj', 'Date MAJ Prix', 'last_api_update', 15),
            ('col_obsolete', 'Obsolète', 'is_obsolete', 10),
            ('col_source', 'Source', 'import_source', 10),
        ]
        
        selected = []
        for field_name, header, data_field, width in all_columns:
            if getattr(self, field_name, False):
                selected.append((header, data_field, width))
        
        return selected

    def _create_rexel_excel(self, articles):
        """Créer le fichier Excel au format Rexel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Articles Rexel"
        
        # Récupérer les colonnes sélectionnées
        columns = self._get_selected_columns()
        
        if not columns:
            raise UserError(_("Veuillez sélectionner au moins une colonne à exporter."))
        
        # Style pour l'en-tête
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Écrire l'en-tête
        for col_idx, (header, _, width) in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[col_letter].width = width
        
        # Figer la première ligne
        ws.freeze_panes = 'A2'
        
        # Écrire les données
        row = 2
        for article in articles:
            # Préparer les données de l'article
            article_data = self._get_article_data(article)
            
            # Écrire chaque colonne sélectionnée
            for col_idx, (_, data_field, _) in enumerate(columns, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = article_data.get(data_field, '')
                cell.border = thin_border
            
            row += 1
        
        # Sauvegarder dans un buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()

    def _get_article_data(self, article):
        """Prépare les données d'un article pour l'export"""
        # Récupérer les infos famille depuis les champs texte ou la relation
        famille_name = article.famille_libelle or ''
        sous_famille_name = article.sous_famille_libelle or ''
        fonction_name = article.fonction_libelle or ''
        famille_code = article.famille_code or ''
        sous_famille_code = article.sous_famille_code or ''
        fonction_code = article.fonction_code or ''
        
        # Si family_node_id existe, utiliser les infos de la hiérarchie
        if article.family_node_id:
            famille = article.family_node_id
            if famille.level == 'fonction':
                fonction_name = famille.name
                fonction_code = famille.code or ''
                if famille.parent_id:
                    sous_famille_name = famille.parent_id.name
                    sous_famille_code = famille.parent_id.code or ''
                    if famille.parent_id.parent_id:
                        famille_name = famille.parent_id.parent_id.name
                        famille_code = famille.parent_id.parent_id.code or ''
            elif famille.level == 'sous_famille':
                sous_famille_name = famille.name
                sous_famille_code = famille.code or ''
                if famille.parent_id:
                    famille_name = famille.parent_id.name
                    famille_code = famille.parent_id.code or ''
            else:
                famille_name = famille.name
                famille_code = famille.code or ''
        
        return {
            'reference_fabricant': article.reference_fabricant or '',
            'trigramme_fabricant': article.trigramme_fabricant or '',
            'fabricant_libelle': article.fabricant_libelle or '',
            'reference_rexel': article.reference_rexel or '',
            'designation': article.designation or '',
            'prix_base': article.prix_base or 0,
            'prix_net': article.prix_net or 0,
            'remise': article.remise or 0,
            'unite_mesure': article.unite_mesure or '',
            'conditionnement': article.conditionnement or '',
            'code_ean13': article.code_ean13 or '',
            'montant_d3e': article.montant_d3e or 0,
            'unite_d3e': article.unite_d3e or 0,
            'famille_libelle': famille_name,
            'sous_famille_libelle': sous_famille_name,
            'fonction_libelle': fonction_name,
            'famille_code': famille_code,
            'sous_famille_code': sous_famille_code,
            'fonction_code': fonction_code,
            'last_api_update': article.last_api_update.strftime('%d/%m/%Y') if article.last_api_update else '',
            'is_obsolete': 'Oui' if article.is_obsolete else 'Non',
            'import_source': article.import_source or '',
        }

    def action_back(self):
        """Retour à la configuration"""
        self.write({'state': 'draft'})
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.rexel.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
