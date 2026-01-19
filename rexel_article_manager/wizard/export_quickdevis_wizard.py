# -*- coding: utf-8 -*-

import base64
import io
from datetime import datetime
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import logging

_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    _logger.warning("openpyxl n'est pas installé. L'export Excel ne sera pas disponible.")


class ExportQuickdevisWizard(models.TransientModel):
    _name = 'export.quickdevis.wizard'
    _description = 'Assistant d\'export vers QuickDevis 7'

    article_ids = fields.Many2many('rexel.article', string='Articles à exporter')
    
    # Options d'export
    export_all = fields.Boolean(string='Exporter tous les articles', default=False)
    include_obsolete = fields.Boolean(string='Inclure les articles obsolètes', default=False)
    
    # Filtres
    famille_ids = fields.Many2many(
        'rexel.product.family',
        'export_qdv_wizard_famille_rel',
        'wizard_id', 'famille_id',
        string='Familles',
        domain=[('parent_id', '=', False)]
    )
    
    # Résultat
    export_file = fields.Binary(string='Fichier exporté', readonly=True)
    export_filename = fields.Char(string='Nom du fichier', readonly=True)
    state = fields.Selection([
        ('draft', 'Configuration'),
        ('done', 'Terminé')
    ], default='draft')
    articles_exported = fields.Integer(string='Articles exportés', readonly=True)

    @api.model
    def default_get(self, fields_list):
        """Récupère les articles sélectionnés"""
        res = super().default_get(fields_list)
        
        if self.env.context.get('active_ids'):
            res['article_ids'] = [(6, 0, self.env.context.get('active_ids'))]
        
        return res

    def _get_all_subfamilies(self, family_ids):
        """Récupère récursivement toutes les sous-familles"""
        Family = self.env['rexel.product.family']
        all_ids = list(family_ids)
        children = Family.search([('parent_id', 'in', family_ids)])
        if children:
            all_ids.extend(self._get_all_subfamilies(children.ids))
        return all_ids

    def action_export(self):
        """Lance l'export vers QuickDevis 7 avec le format exact"""
        self.ensure_one()
        
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("La bibliothèque openpyxl n'est pas installée."))
        
        # Déterminer les articles à exporter
        if self.export_all:
            domain = []
            if not self.include_obsolete:
                domain.append(('is_obsolete', '=', False))
            articles = self.env['rexel.article'].search(domain)
        elif self.article_ids:
            articles = self.article_ids
            if not self.include_obsolete:
                articles = articles.filtered(lambda a: not a.is_obsolete)
        elif self.famille_ids:
            all_family_ids = self._get_all_subfamilies(self.famille_ids.ids)
            domain = [('family_node_id', 'in', all_family_ids)]
            if not self.include_obsolete:
                domain.append(('is_obsolete', '=', False))
            articles = self.env['rexel.article'].search(domain)
        else:
            raise UserError(_("Veuillez sélectionner des articles ou cocher 'Exporter tous les articles'."))
        
        if not articles:
            raise UserError(_('Aucun article à exporter.'))
        
        # Créer le fichier Excel
        excel_data = self._create_quickdevis_excel(articles)
        
        # Préparer le nom du fichier
        filename = f"Export_QuickDevis7_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        # Sauvegarder
        self.write({
            'export_file': base64.b64encode(excel_data),
            'export_filename': filename,
            'state': 'done',
            'articles_exported': len(articles),
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.quickdevis.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _create_quickdevis_excel(self, articles):
        """
        Créer le fichier Excel au format QuickDevis 7 EXACT
        
        Structure onglet "Articles Rexel":
        - Ligne 1: En-têtes QDV7 (mapping pour import)
        - Ligne 2: Formule =...
        - Ligne 3: En-têtes descriptifs
        - Ligne 4+: Données avec formules
        
        Structure onglet "Structure":
        - Ligne 1: En-têtes QDV7
        - Ligne 2: Formule =...
        - Ligne 3+: Familles puis Sous-familles
        """
        wb = Workbook()
        
        # ==========================================
        # ONGLET 1 : Articles Rexel
        # ==========================================
        ws_articles = wb.active
        ws_articles.title = "Articles Rexel"
        
        # Styles
        header_font = Font(bold=True)
        
        # ===== LIGNE 1 : En-têtes QDV7 (mapping) =====
        qdv7_headers = {
            'A': '*',
            'B': 'Référence',
            'E': 'Description',
            'G': 'MATERIAL / Coût unitaire',
            'H': 'MATERIAL / Remise',
            'I': 'Unité',
            'P': 'Description structure',
            'T': 'Date de modification',
            'V': 'Code structure',
            'W': 'Fabricant',
            'X': 'Famille',
        }
        for col_letter, header in qdv7_headers.items():
            ws_articles[f'{col_letter}1'] = header
            ws_articles[f'{col_letter}1'].font = header_font
        
        # ===== LIGNE 2 : Marqueur spécial QDV7 (texte, pas formule) =====
        # Préfixer avec apostrophe pour forcer le texte dans Excel
        ws_articles['A2'] = "'=..."
        ws_articles['A2'].data_type = 's'  # Forcer le type string
        
        # ===== LIGNE 3 : En-têtes descriptifs =====
        descriptive_headers = {
            'B': 'Référence Fabricant',
            'C': 'Trigramme Fabricant',
            'D': 'Référence Rexel',
            'E': 'Désignation',
            'F': 'Prix Base',
            'G': 'Prix Net',
            'H': 'Remise %',
            'I': 'Unité Mesure',
            'J': 'Conditionnement',
            'K': 'Code EAN',
            'L': 'Montant D3E',
            'M': 'Unité D3E',
            'N': 'Famille',
            'O': 'Sous-Famille',
            'P': 'Fonction',
            'Q': 'Code Famille',
            'R': 'Code Sous-Famille',
            'S': 'Code Fonction',
            'T': 'Date MAJ Prix',
            'U': 'Obsolète',
            'V': 'Code structure',
            'W': 'Fabriquant',
            'X': 'Code complet famille',
        }
        for col_letter, header in descriptive_headers.items():
            ws_articles[f'{col_letter}3'] = header
            ws_articles[f'{col_letter}3'].font = header_font
        
        # Largeurs de colonnes
        column_widths = {
            'A': 3, 'B': 20, 'C': 15, 'D': 20, 'E': 50, 'F': 12, 'G': 12,
            'H': 10, 'I': 12, 'J': 15, 'K': 15, 'L': 12, 'M': 10,
            'N': 35, 'O': 35, 'P': 35, 'Q': 12, 'R': 15, 'S': 12,
            'T': 15, 'U': 10, 'V': 15, 'W': 25, 'X': 20,
        }
        for col_letter, width in column_widths.items():
            ws_articles.column_dimensions[col_letter].width = width
        
        # Figer les 3 premières lignes
        ws_articles.freeze_panes = 'A4'
        
        # Collecter les données pour l'onglet Structure
        familles_dict = {}  # code_famille -> libelle_famille
        sous_familles_dict = {}  # (code_famille + code_sous_famille) -> libelle_sous_famille
        
        # ===== LIGNE 4+ : Données =====
        row = 4
        for article in articles:
            data = self._get_article_data(article)
            
            # Collecter pour l'onglet Structure
            if data['famille_code'] and data['famille_libelle']:
                familles_dict[data['famille_code']] = data['famille_libelle']
            if data['famille_code'] and data['sous_famille_code'] and data['sous_famille_libelle']:
                code_complet = f"{data['famille_code']}{data['sous_famille_code']}"
                sous_familles_dict[code_complet] = data['sous_famille_libelle']
            
            # Écrire les données
            ws_articles[f'B{row}'] = data['reference_fabricant']
            ws_articles[f'C{row}'] = data['trigramme_fabricant']
            ws_articles[f'D{row}'] = data['reference_rexel']
            ws_articles[f'E{row}'] = data['designation']
            ws_articles[f'F{row}'] = data['prix_base']
            ws_articles[f'G{row}'] = data['prix_net']
            ws_articles[f'H{row}'] = data['remise']
            ws_articles[f'I{row}'] = data['unite_mesure']
            ws_articles[f'J{row}'] = data['conditionnement']
            ws_articles[f'K{row}'] = data['code_ean13']
            ws_articles[f'L{row}'] = data['montant_d3e']
            ws_articles[f'M{row}'] = data['unite_d3e']
            ws_articles[f'N{row}'] = data['famille_libelle']
            ws_articles[f'O{row}'] = data['sous_famille_libelle']
            ws_articles[f'P{row}'] = data['fonction_libelle']
            ws_articles[f'Q{row}'] = data['famille_code']
            ws_articles[f'R{row}'] = data['sous_famille_code']
            ws_articles[f'S{row}'] = data['fonction_code']
            ws_articles[f'T{row}'] = data['date_maj']
            ws_articles[f'U{row}'] = data['is_obsolete']
            
            # Colonne V : Formule Code structure =Q&R&S
            ws_articles[f'V{row}'] = f'=Q{row}&R{row}&S{row}'
            
            # Colonne W : Fabricant (libellé)
            ws_articles[f'W{row}'] = data['fabricant_libelle']
            
            # Colonne X : Formule Code complet famille =Q&R&S
            ws_articles[f'X{row}'] = f'=Q{row}&R{row}&S{row}'
            
            row += 1
        
        # ==========================================
        # ONGLET 2 : Structure
        # ==========================================
        ws_structure = wb.create_sheet(title="Structure")
        
        # ===== LIGNE 1 : En-têtes QDV7 =====
        ws_structure['A1'] = '*'
        ws_structure['A1'].font = header_font
        ws_structure['B1'] = 'Description structure'
        ws_structure['B1'].font = header_font
        ws_structure['C1'] = 'Code structure'
        ws_structure['C1'].font = header_font
        
        # ===== LIGNE 2 : Marqueur spécial QDV7 (texte, pas formule) =====
        ws_structure['A2'] = "'=..."
        ws_structure['A2'].data_type = 's'  # Forcer le type string
        
        # Largeurs de colonnes
        ws_structure.column_dimensions['A'].width = 3
        ws_structure.column_dimensions['B'].width = 50
        ws_structure.column_dimensions['C'].width = 20
        
        # Figer les 2 premières lignes
        ws_structure.freeze_panes = 'A3'
        
        # ===== LIGNE 3+ : Familles puis Sous-familles =====
        struct_row = 3
        
        # D'abord les FAMILLES (triées par code)
        for code in sorted(familles_dict.keys()):
            libelle = familles_dict[code]
            ws_structure[f'B{struct_row}'] = libelle
            ws_structure[f'C{struct_row}'] = code
            struct_row += 1
        
        # Ensuite les SOUS-FAMILLES (triées par code)
        for code in sorted(sous_familles_dict.keys()):
            libelle = sous_familles_dict[code]
            ws_structure[f'B{struct_row}'] = libelle
            ws_structure[f'C{struct_row}'] = code
            struct_row += 1
        
        # ==========================================
        # ONGLET 3 : _qdv_settings (optionnel, pour compatibilité)
        # ==========================================
        ws_settings = wb.create_sheet(title="_qdv_settings")
        ws_settings['A1'] = 'Version'
        ws_settings['B1'] = '7.0'
        
        # Sauvegarder dans un buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()

    def _get_article_data(self, article):
        """Prépare les données d'un article pour l'export"""
        # Récupérer les infos famille depuis les champs texte ou la relation
        famille_name = article.famille_libelle or ''
        sous_famille_name = article.sous_famille_libelle or ''
        fonction_name = article.fonction_libelle or ''
        famille_code = article.famille_code or ''
        sous_famille_code = article.sous_famille_code or ''
        fonction_code = article.fonction_code or ''
        
        # Si family_node_id existe, utiliser les infos de la hiérarchie
        if article.family_node_id:
            famille = article.family_node_id
            if famille.level == 'fonction':
                fonction_name = famille.name
                fonction_code = famille.code or ''
                if famille.parent_id:
                    sous_famille_name = famille.parent_id.name
                    sous_famille_code = famille.parent_id.code or ''
                    if famille.parent_id.parent_id:
                        famille_name = famille.parent_id.parent_id.name
                        famille_code = famille.parent_id.parent_id.code or ''
            elif famille.level == 'sous_famille':
                sous_famille_name = famille.name
                sous_famille_code = famille.code or ''
                if famille.parent_id:
                    famille_name = famille.parent_id.name
                    famille_code = famille.parent_id.code or ''
            else:
                famille_name = famille.name
                famille_code = famille.code or ''
        
        # Date de mise à jour
        date_maj = ''
        if article.last_api_update:
            date_maj = article.last_api_update.strftime('%d/%m/%Y')
        elif article.date_tarif:
            date_maj = article.date_tarif.strftime('%d/%m/%Y')
        
        return {
            'reference_fabricant': article.reference_fabricant or '',
            'trigramme_fabricant': article.trigramme_fabricant or '',
            'fabricant_libelle': article.fabricant_libelle or '',
            'reference_rexel': article.reference_rexel or '',
            'designation': article.designation or '',
            'prix_base': article.prix_base or 0,
            'prix_net': article.prix_net or 0,
            'remise': article.remise or 0,
            'unite_mesure': article.unite_mesure or 'U',
            'conditionnement': article.conditionnement or '',
            'code_ean13': article.code_ean13 or '',
            'montant_d3e': article.montant_d3e or 0,
            'unite_d3e': article.unite_d3e or 0,
            'famille_libelle': famille_name,
            'sous_famille_libelle': sous_famille_name,
            'fonction_libelle': fonction_name,
            'famille_code': famille_code,
            'sous_famille_code': sous_famille_code,
            'fonction_code': fonction_code,
            'date_maj': date_maj,
            'is_obsolete': 'Oui' if article.is_obsolete else 'Non',
        }

    def action_back(self):
        """Retour à la configuration"""
        self.write({'state': 'draft'})
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.quickdevis.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_download(self):
        """Télécharge le fichier exporté"""
        self.ensure_one()
        
        if not self.export_file:
            raise UserError(_('Aucun fichier à télécharger. Veuillez d\'abord lancer l\'export.'))
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/export.quickdevis.wizard/{self.id}/export_file/{self.export_filename}?download=true',
            'target': 'self',
        }
